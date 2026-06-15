# ui_main.py
# 主應用視窗
# Main Application Window

from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QTableWidget, QTableWidgetItem, QLineEdit, QComboBox,
    QSpinBox, QDateEdit, QTextEdit, QDialog, QMessageBox,
    QTabWidget, QInputDialog, QFileDialog, QProgressDialog,
    QApplication, QHeaderView, QScrollArea, QCheckBox
)
from PyQt6.QtCore import Qt, QDate, QDateTime, pyqtSignal
from PyQt6.QtGui import QFont, QColor, QIcon
import json
from datetime import datetime, timedelta

from models import DatabaseManager
from services import (
    UserService, ReagentService, StockService, InventoryCheckService,
    ScrapService, AuditLogService, TraceabilityService, QCService
)
from utils import (
    BarcodeGenerator, LabelPrinter, CodeGenerator, DateHelper,
    ValidationHelper, NetworkHelper
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


class MainWindow(QMainWindow):

    def __init__(self, user_id, db_path='reagent_system.db'):
        super().__init__()

        self.db_manager = DatabaseManager(db_path)
        self.session = self.db_manager.get_session()

        # 從新 session 讀取 user，完全避免 DetachedInstanceError
        from models import User
        user = self.session.query(User).filter_by(id=user_id).first()

        self.user_id = user.id
        self.user_real_name = user.real_name
        self.user_role = user.role
        self.user_department = user.department or ''

        # 初始化服務
        self.user_service = UserService(self.session)
        self.reagent_service = ReagentService(self.session)
        self.stock_service = StockService(self.session)
        self.check_service = InventoryCheckService(self.session)
        self.scrap_service = ScrapService(self.session)
        self.audit_service = AuditLogService(self.session)
        self.traceability_service = TraceabilityService(self.session)
        self.qc_service = QCService(self.session)

        self.setWindowTitle(f"醫療檢驗試劑管理系統 - {self.user_real_name} ({self.user_role})")
        self.setGeometry(0, 0, 1920, 1080)
        self.showMaximized()

        self.init_ui()
    
    def init_ui(self):
        """初始化主視窗UI"""
        # 中央widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout()
        
        # 頂部按鈕區
        button_layout = self.create_button_bar()
        main_layout.addLayout(button_layout)
        
        # 主要內容區（tab widget）
        self.tab_widget = QTabWidget()
        main_layout.addWidget(self.tab_widget)
        
        # 建立各個標籤頁
        self.tab_widget.addTab(self.create_dashboard_tab(), "首頁")
        self.tab_widget.addTab(self.create_reagent_tab(), "試劑設定")
        self.tab_widget.addTab(self.create_stock_in_tab(), "入庫")
        self.tab_widget.addTab(self.create_stock_out_tab(), "出庫")
        self.tab_widget.addTab(self.create_inventory_tab(), "查詢")
        self.tab_widget.addTab(self.create_qc_tab(), "驗收")
        self.tab_widget.addTab(self.create_check_tab(), "盤點")
        self.tab_widget.addTab(self.create_scrap_tab(), "報廢")
        self.tab_widget.addTab(self.create_traceability_tab(), "追溯")
        
        if self.user_role == 'admin':
            self.tab_widget.addTab(self.create_report_tab(), "報表")
            self.tab_widget.addTab(self.create_admin_tab(), "系統設定")
        
        central_widget.setLayout(main_layout)
    
    def create_button_bar(self):
        """建立頂部按鈕欄"""
        layout = QHBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        buttons_config = [
            ("試劑設定", 1),
            ("入庫", 2),
            ("出庫", 3),
            ("盤點", 5),
            ("查詢", 4),
            ("報廢", 6),
        ]
        
        for btn_text, tab_index in buttons_config:
            btn = QPushButton(btn_text)
            btn.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            btn.setMinimumHeight(50)
            btn.setMinimumWidth(120)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #0078D4;
                    color: white;
                    border: none;
                    border-radius: 5px;
                }
                QPushButton:hover {
                    background-color: #005A9E;
                }
            """)
            btn.clicked.connect(lambda checked, idx=tab_index: self.tab_widget.setCurrentIndex(idx))
            layout.addWidget(btn)
        
        # 登出按鈕
        layout.addStretch()
        logout_btn = QPushButton("登出")
        logout_btn.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        logout_btn.setMinimumHeight(50)
        logout_btn.setMinimumWidth(120)
        logout_btn.setStyleSheet("""
            QPushButton {
                background-color: #D13438;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #A4373A;
            }
        """)
        logout_btn.clicked.connect(self.logout)
        layout.addWidget(logout_btn)
        
        return layout
    
    def create_dashboard_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("庫存管理儀表板")
        title.setFont(QFont("Arial", 16, QFont.Weight.Bold))
        layout.addWidget(title)

        # 統計卡片（用 QLabel 儲存參考以便更新）
        stats_layout = QHBoxLayout()
        self._stat_cards = {}
        for key, label, color in [
            ('total',    '庫存總數',   '#0078D4'),
            ('in_today', '今日入庫',   '#107C10'),
            ('out_today','今日出庫',   '#FFB900'),
            ('expiring', '即將到期',   '#D13438'),
            ('expired',  '已過期',     '#8B0000'),
            ('low',      '低於安全庫存','#FF8C00'),
        ]:
            card = QWidget()
            card.setStyleSheet(f"QWidget{{background:{color};border-radius:6px;padding:12px;}}")
            cl = QVBoxLayout(card)
            lbl = QLabel(label)
            lbl.setStyleSheet("color:white;font-size:12px;")
            val = QLabel("0")
            val.setStyleSheet("color:white;font-size:28px;font-weight:bold;")
            cl.addWidget(lbl)
            cl.addWidget(val)
            self._stat_cards[key] = val
            stats_layout.addWidget(card)
        layout.addLayout(stats_layout)
        layout.addSpacing(12)

        expiry_label = QLabel("快過期試劑（90天內）")
        expiry_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(expiry_label)
        self.expiry_table = QTableWidget()
        self.expiry_table.setColumnCount(5)
        self.expiry_table.setHorizontalHeaderLabels(["試劑名稱","LOT號","有效期限","剩餘天數","庫存"])
        self.expiry_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.expiry_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.expiry_table)

        low_label = QLabel("低庫存警示")
        low_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(low_label)
        self.low_stock_table = QTableWidget()
        self.low_stock_table.setColumnCount(4)
        self.low_stock_table.setHorizontalHeaderLabels(["試劑名稱","現有總庫存","安全庫存","缺少"])
        self.low_stock_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.low_stock_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.low_stock_table)

        layout.addStretch()
        self.refresh_dashboard()
        widget.setLayout(layout)
        return widget
    
    def create_reagent_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("試劑設定")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        btn_layout = QHBoxLayout()
        add_btn = QPushButton("新增試劑")
        add_btn.setMinimumHeight(40)
        add_btn.clicked.connect(self.add_reagent_dialog)
        btn_layout.addWidget(add_btn)

        edit_btn = QPushButton("編輯選取")
        edit_btn.setMinimumHeight(40)
        edit_btn.clicked.connect(self.edit_reagent_dialog)
        btn_layout.addWidget(edit_btn)

        delete_btn = QPushButton("刪除")
        delete_btn.setMinimumHeight(40)
        delete_btn.setStyleSheet("background-color: #D13438; color: white;")
        delete_btn.clicked.connect(self.delete_reagent)
        btn_layout.addWidget(delete_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.setMinimumHeight(40)
        refresh_btn.clicked.connect(self.refresh_reagent_table)
        btn_layout.addWidget(refresh_btn)

        download_btn = QPushButton("下載匯入範本")
        download_btn.setMinimumHeight(40)
        download_btn.setStyleSheet("background-color: #0078D4; color: white;")
        download_btn.clicked.connect(self.download_reagent_template)
        btn_layout.addWidget(download_btn)

        import_btn = QPushButton("匯入 Excel")
        import_btn.setMinimumHeight(40)
        import_btn.setStyleSheet("background-color: #107C10; color: white;")
        import_btn.clicked.connect(self.import_reagent_excel)
        btn_layout.addWidget(import_btn)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        self.reagent_table = QTableWidget()
        self.reagent_table.setColumnCount(14)
        self.reagent_table.setHorizontalHeaderLabels([
            "ID", "條碼前16碼", "名稱", "廠牌", "供應商",
            "規格", "單位", "安全庫存",
            "批號起始", "批號長度", "效期起始", "效期長度", "效期格式", "保管人"
        ])
        self.reagent_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.reagent_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.reagent_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.reagent_table)

        self.refresh_reagent_table()
        widget.setLayout(layout)
        return widget
    
    def create_stock_in_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(12)

        title = QLabel("入庫（掃描條碼）")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # 掃描條碼 - 主要輸入
        scan_layout = QHBoxLayout()
        scan_layout.addWidget(QLabel("掃描條碼:"))
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("掃描試劑條碼，自動識別試劑、批號、效期")
        self.barcode_input.setMinimumHeight(40)
        self.barcode_input.setStyleSheet("font-size: 14px;")
        self.barcode_input.returnPressed.connect(self.parse_scanned_barcode)
        scan_layout.addWidget(self.barcode_input)

        parse_btn = QPushButton("解析")
        parse_btn.setMinimumHeight(40)
        parse_btn.setMinimumWidth(80)
        parse_btn.clicked.connect(self.parse_scanned_barcode)
        scan_layout.addWidget(parse_btn)

        search_btn = QPushButton("🔍 搜尋試劑")
        search_btn.setMinimumHeight(40)
        search_btn.setMinimumWidth(100)
        search_btn.setStyleSheet("background-color: #FF8C00; color: white; border-radius: 4px; font-weight: bold;")
        search_btn.clicked.connect(self.open_reagent_search_dialog)
        scan_layout.addWidget(search_btn)
        layout.addLayout(scan_layout)

        # 解析後的資訊（唯讀顯示）
        info_group = QWidget()
        info_group.setStyleSheet("QWidget { background: #F5F5F5; border-radius: 5px; padding: 8px; }")
        info_layout = QHBoxLayout(info_group)

        info_layout.addWidget(QLabel("試劑名稱:"))
        self.stock_in_reagent_name = QLineEdit()
        self.stock_in_reagent_name.setReadOnly(True)
        self.stock_in_reagent_name.setStyleSheet("background: white;")
        info_layout.addWidget(self.stock_in_reagent_name)

        info_layout.addWidget(QLabel("LOT號:"))
        self.stock_in_lot = QLineEdit()
        self.stock_in_lot.setMinimumHeight(36)
        info_layout.addWidget(self.stock_in_lot)

        info_layout.addWidget(QLabel("有效期限:"))
        self.stock_in_expiry = QLineEdit()
        self.stock_in_expiry.setPlaceholderText("YYYY-MM-DD")
        self.stock_in_expiry.setMinimumHeight(36)
        info_layout.addWidget(self.stock_in_expiry)

        layout.addWidget(info_group)

        # 數量與備註
        form_layout = QHBoxLayout()
        form_layout.addWidget(QLabel("入庫盒數:"))
        self.stock_in_qty = QSpinBox()
        self.stock_in_qty.setMinimum(1)
        self.stock_in_qty.setMaximum(99999)
        self.stock_in_qty.setValue(1)
        self.stock_in_qty.setMinimumHeight(36)
        # Enter 鍵觸發入庫
        self.stock_in_qty.lineEdit().returnPressed.connect(self.do_stock_in)
        form_layout.addWidget(self.stock_in_qty)

        self.stock_in_box_hint = QLabel("")
        self.stock_in_box_hint.setStyleSheet("color: #0078D4; font-weight: bold;")
        form_layout.addWidget(self.stock_in_box_hint)

        form_layout.addWidget(QLabel("採購單號:"))
        self.stock_in_po = QLineEdit()
        self.stock_in_po.setMinimumHeight(36)
        form_layout.addWidget(self.stock_in_po)

        form_layout.addWidget(QLabel("備註:"))
        self.stock_in_remark = QLineEdit()
        self.stock_in_remark.setMinimumHeight(36)
        form_layout.addWidget(self.stock_in_remark)
        layout.addLayout(form_layout)

        # 按鈕
        btn_layout = QHBoxLayout()
        stock_in_btn = QPushButton("確認入庫")
        stock_in_btn.setMinimumHeight(45)
        stock_in_btn.setMinimumWidth(120)
        stock_in_btn.setStyleSheet("QPushButton { background-color: #107C10; color: white; font-weight: bold; font-size: 13px; border-radius: 5px; } QPushButton:hover { background-color: #0B6104; }")
        stock_in_btn.clicked.connect(self.do_stock_in)
        btn_layout.addWidget(stock_in_btn)

        clear_btn = QPushButton("清除")
        clear_btn.setMinimumHeight(45)
        clear_btn.setMinimumWidth(80)
        clear_btn.clicked.connect(self.clear_stock_in_form)
        btn_layout.addWidget(clear_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 今日入庫記錄
        today_label = QLabel("今日入庫記錄")
        today_label.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        layout.addWidget(today_label)

        self.stock_in_history = QTableWidget()
        self.stock_in_history.setColumnCount(7)
        self.stock_in_history.setHorizontalHeaderLabels(["時間", "試劑名稱", "LOT號", "效期", "數量", "院內編號", "補列印"])
        self.stock_in_history.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.stock_in_history.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.stock_in_history)

        # 隱藏的試劑ID
        self.stock_in_reagent_id = None

        widget.setLayout(layout)
        return widget
    
    def create_stock_out_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(12)

        title = QLabel("出庫")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # 掃描院內編號
        scan_layout = QHBoxLayout()
        scan_layout.addWidget(QLabel("掃描院內編號:"))
        self.stock_out_barcode = QLineEdit()
        self.stock_out_barcode.setPlaceholderText("掃描或輸入院內編號 R202606XXXX")
        self.stock_out_barcode.setMinimumHeight(40)
        self.stock_out_barcode.setStyleSheet("font-size: 14px;")
        self.stock_out_barcode.returnPressed.connect(self.scan_stock_out_barcode)
        scan_layout.addWidget(self.stock_out_barcode)

        out_search_btn = QPushButton("🔍 搜尋庫存")
        out_search_btn.setMinimumHeight(40)
        out_search_btn.setMinimumWidth(100)
        out_search_btn.setStyleSheet("background-color: #FF8C00; color: white; border-radius: 4px; font-weight: bold;")
        out_search_btn.clicked.connect(self.open_inventory_search_dialog)
        scan_layout.addWidget(out_search_btn)

        layout.addLayout(scan_layout)

        # 試劑資訊
        info_group = QWidget()
        info_group.setStyleSheet("QWidget { background: #F5F5F5; border-radius: 5px; padding: 8px; }")
        info_layout = QHBoxLayout(info_group)

        info_layout.addWidget(QLabel("試劑名稱:"))
        self.stock_out_name = QLineEdit()
        self.stock_out_name.setReadOnly(True)
        self.stock_out_name.setStyleSheet("background: white;")
        info_layout.addWidget(self.stock_out_name)

        info_layout.addWidget(QLabel("LOT號:"))
        self.stock_out_lot = QLineEdit()
        self.stock_out_lot.setReadOnly(True)
        self.stock_out_lot.setStyleSheet("background: white;")
        info_layout.addWidget(self.stock_out_lot)

        info_layout.addWidget(QLabel("效期:"))
        self.stock_out_expiry = QLineEdit()
        self.stock_out_expiry.setReadOnly(True)
        self.stock_out_expiry.setStyleSheet("background: white;")
        info_layout.addWidget(self.stock_out_expiry)

        info_layout.addWidget(QLabel("現有庫存:"))
        self.stock_out_available = QLineEdit()
        self.stock_out_available.setReadOnly(True)
        self.stock_out_available.setStyleSheet("background: white;")
        info_layout.addWidget(self.stock_out_available)

        layout.addWidget(info_group)

        # 出庫表單
        form_layout = QHBoxLayout()
        form_layout.addWidget(QLabel("出庫數量:"))
        self.stock_out_qty = QSpinBox()
        self.stock_out_qty.setMinimum(1)
        self.stock_out_qty.setMaximum(99999)
        self.stock_out_qty.setMinimumHeight(36)
        form_layout.addWidget(self.stock_out_qty)

        form_layout.addWidget(QLabel("使用單位:"))
        self.stock_out_dept = QLineEdit()
        self.stock_out_dept.setMinimumHeight(36)
        form_layout.addWidget(self.stock_out_dept)

        form_layout.addWidget(QLabel("使用儀器:"))
        self.stock_out_equipment = QLineEdit()
        self.stock_out_equipment.setMinimumHeight(36)
        form_layout.addWidget(self.stock_out_equipment)

        form_layout.addWidget(QLabel("備註:"))
        self.stock_out_remark = QLineEdit()
        self.stock_out_remark.setMinimumHeight(36)
        form_layout.addWidget(self.stock_out_remark)
        layout.addLayout(form_layout)

        # 按鈕
        btn_layout = QHBoxLayout()
        stock_out_btn = QPushButton("確認出庫")
        stock_out_btn.setMinimumHeight(45)
        stock_out_btn.setMinimumWidth(120)
        stock_out_btn.setStyleSheet("QPushButton { background-color: #FF8C00; color: white; font-weight: bold; font-size: 13px; border-radius: 5px; } QPushButton:hover { background-color: #E07800; }")
        stock_out_btn.clicked.connect(self.do_stock_out)
        btn_layout.addWidget(stock_out_btn)

        clear_btn = QPushButton("清除")
        clear_btn.setMinimumHeight(45)
        clear_btn.setMinimumWidth(80)
        clear_btn.clicked.connect(self.clear_stock_out_form)
        btn_layout.addWidget(clear_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # 今日出庫記錄（含補列印）
        out_label = QLabel("今日出庫記錄")
        out_label.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        layout.addWidget(out_label)

        self.stock_out_history = QTableWidget()
        self.stock_out_history.setColumnCount(7)
        self.stock_out_history.setHorizontalHeaderLabels(
            ["時間", "試劑名稱", "LOT號", "效期", "出庫數量", "院內編號", "補列印"])
        self.stock_out_history.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.stock_out_history.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.stock_out_history)

        # 保存 stock_in 資訊供出庫用
        self._current_stock_in = None

        widget.setLayout(layout)
        return widget
    
    def create_inventory_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)

        title = QLabel("查詢")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # 查詢類型切換
        type_layout = QHBoxLayout()
        self.query_type = QComboBox()
        self.query_type.addItems(["目前庫存", "入庫紀錄", "出庫紀錄"])
        self.query_type.setMinimumHeight(36)
        self.query_type.currentIndexChanged.connect(self._toggle_query_ui)
        type_layout.addWidget(QLabel("查詢類型:"))
        type_layout.addWidget(self.query_type)
        type_layout.addStretch()
        layout.addLayout(type_layout)

        # 查詢條件
        cond_layout = QHBoxLayout()

        cond_layout.addWidget(QLabel("試劑名稱:"))
        self.query_name = QLineEdit()
        self.query_name.setMinimumHeight(34)
        cond_layout.addWidget(self.query_name)

        cond_layout.addWidget(QLabel("LOT號:"))
        self.query_lot = QLineEdit()
        self.query_lot.setMinimumHeight(34)
        cond_layout.addWidget(self.query_lot)

        # 日期區間（入出庫用）
        self.query_date_from_label = QLabel("開始日期:")
        cond_layout.addWidget(self.query_date_from_label)
        self.query_date_from = QDateEdit()
        self.query_date_from.setDate(QDate.currentDate().addDays(-30))
        self.query_date_from.setCalendarPopup(True)
        self.query_date_from.setMinimumHeight(34)
        cond_layout.addWidget(self.query_date_from)

        self.query_date_to_label = QLabel("結束日期:")
        cond_layout.addWidget(self.query_date_to_label)
        self.query_date_to = QDateEdit()
        self.query_date_to.setDate(QDate.currentDate())
        self.query_date_to.setCalendarPopup(True)
        self.query_date_to.setMinimumHeight(34)
        cond_layout.addWidget(self.query_date_to)

        search_btn = QPushButton("查詢")
        search_btn.setMinimumHeight(36)
        search_btn.setMinimumWidth(80)
        search_btn.setStyleSheet("background-color: #0078D4; color: white; border-radius: 4px;")
        search_btn.clicked.connect(self.query_inventory)
        cond_layout.addWidget(search_btn)

        clear_btn = QPushButton("清除")
        clear_btn.setMinimumHeight(36)
        clear_btn.clicked.connect(self.clear_query_conditions)
        cond_layout.addWidget(clear_btn)

        layout.addLayout(cond_layout)

        # 結果筆數
        self.query_count_label = QLabel("共 0 筆")
        layout.addWidget(self.query_count_label)

        # 查詢結果表格
        self.inventory_table = QTableWidget()
        self.inventory_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.inventory_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.inventory_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.inventory_table)

        # 預設隱藏日期欄位（庫存模式不需要）
        self._toggle_query_ui(0)

        widget.setLayout(layout)
        return widget

    def _toggle_query_ui(self, idx):
        """切換查詢類型時顯示/隱藏日期欄位"""
        show_date = (idx != 0)
        self.query_date_from_label.setVisible(show_date)
        self.query_date_from.setVisible(show_date)
        self.query_date_to_label.setVisible(show_date)
        self.query_date_to.setVisible(show_date)

    def query_inventory(self):
        try:
            self.inventory_table.setRowCount(0)
            idx = self.query_type.currentIndex()
            name_filter = self.query_name.text().strip()
            lot_filter  = self.query_lot.text().strip()

            if idx == 0:
                self._query_stock(name_filter, lot_filter)
            elif idx == 1:
                date_from = self.query_date_from.date().toPyDate()
                date_to   = self.query_date_to.date().toPyDate()
                self._query_stock_in(name_filter, lot_filter, date_from, date_to)
            else:
                date_from = self.query_date_from.date().toPyDate()
                date_to   = self.query_date_to.date().toPyDate()
                self._query_stock_out(name_filter, lot_filter, date_from, date_to)

        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def _query_stock(self, name_filter, lot_filter):
        """目前庫存"""
        from models import Inventory, ReagentMaster
        self.inventory_table.setColumnCount(7)
        self.inventory_table.setHorizontalHeaderLabels(
            ["試劑名稱", "LOT號", "有效期限", "院內編號", "現有庫存", "安全庫存", "狀態"])

        q = self.session.query(Inventory).join(ReagentMaster)
        if name_filter:
            q = q.filter(ReagentMaster.reagent_name.contains(name_filter))
        if lot_filter:
            q = q.filter(Inventory.lot_number.contains(lot_filter))
        items = q.filter(Inventory.current_quantity > 0).all()

        for idx, item in enumerate(items):
            self.inventory_table.insertRow(idx)
            status, color = DateHelper.get_expiry_status(item.expiry_date)
            vals = [item.reagent.reagent_name, item.lot_number, item.expiry_date,
                    item.in_house_code or '', str(item.current_quantity),
                    str(item.reagent.safety_stock), status]
            for col, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                if col == 6:
                    cell.setBackground(QColor(color))
                    cell.setForeground(QColor('white'))
                self.inventory_table.setItem(idx, col, cell)
        self.query_count_label.setText(f"共 {len(items)} 筆")

    def _query_stock_in(self, name_filter, lot_filter, date_from, date_to):
        from models import StockIn, ReagentMaster
        from datetime import datetime as dt
        self.inventory_table.setColumnCount(8)
        self.inventory_table.setHorizontalHeaderLabels(
            ["入庫日期", "試劑名稱", "LOT號", "有效期限", "數量", "院內編號", "採購單號", "重印"])

        q = self.session.query(StockIn).join(
            ReagentMaster, StockIn.reagent_id == ReagentMaster.id)
        if name_filter:
            q = q.filter(ReagentMaster.reagent_name.contains(name_filter))
        if lot_filter:
            q = q.filter(StockIn.lot_number.contains(lot_filter))
        q = q.filter(
            StockIn.stock_in_date >= dt.combine(date_from, dt.min.time()),
            StockIn.stock_in_date <= dt.combine(date_to,   dt.max.time())
        ).order_by(StockIn.stock_in_date.desc())
        items = q.all()

        for idx, item in enumerate(items):
            self.inventory_table.insertRow(idx)
            vals = [item.stock_in_date.strftime('%Y-%m-%d %H:%M'),
                    item.reagent.reagent_name, item.lot_number,
                    item.expiry_date, str(item.quantity),
                    item.in_house_code or '', item.po_number or '']
            for col, v in enumerate(vals):
                self.inventory_table.setItem(idx, col, QTableWidgetItem(v))
            # 重印按鈕
            code = item.in_house_code or ''
            btn = QPushButton("重印")
            btn.setStyleSheet("background-color: #107C10; color: white; border-radius: 3px;")
            rname = item.reagent.reagent_name
            btn.clicked.connect(lambda _, c=code, n=rname:
                                self._print_label(n, '', '', 0, c, label_type='in',
                                                  handler_name=self.user_real_name))
            self.inventory_table.setCellWidget(idx, 7, btn)
        self.query_count_label.setText(f"共 {len(items)} 筆")

    def _query_stock_out(self, name_filter, lot_filter, date_from, date_to):
        from models import StockOut, StockIn, ReagentMaster, User
        from datetime import datetime as dt
        self.inventory_table.setColumnCount(8)
        self.inventory_table.setHorizontalHeaderLabels(
            ["出庫日期", "試劑名稱", "LOT號", "效期", "出庫數量", "使用單位", "出庫人", "重印"])

        q = (self.session.query(StockOut)
             .join(StockIn, StockOut.stock_in_id == StockIn.id)
             .join(ReagentMaster, StockIn.reagent_id == ReagentMaster.id))
        if name_filter:
            q = q.filter(ReagentMaster.reagent_name.contains(name_filter))
        if lot_filter:
            q = q.filter(StockIn.lot_number.contains(lot_filter))
        q = q.filter(
            StockOut.stock_out_date >= dt.combine(date_from, dt.min.time()),
            StockOut.stock_out_date <= dt.combine(date_to,   dt.max.time())
        ).order_by(StockOut.stock_out_date.desc())
        items = q.all()

        for idx, item in enumerate(items):
            self.inventory_table.insertRow(idx)
            # 取出庫人姓名
            handler = (self.session.query(User).filter_by(id=item.handler_id).first()
                       if item.handler_id else None)
            handler_name = handler.real_name if handler else ''
            rname = item.stock_in.reagent.reagent_name
            exp   = item.stock_in.expiry_date
            code  = item.stock_in.in_house_code or ''
            vals = [item.stock_out_date.strftime('%Y-%m-%d %H:%M'),
                    rname, item.stock_in.lot_number, exp,
                    str(item.quantity), item.usage_department or '',
                    handler_name]
            for col, v in enumerate(vals):
                self.inventory_table.setItem(idx, col, QTableWidgetItem(v))
            # 重印按鈕
            btn = QPushButton("重印")
            btn.setStyleSheet("background-color: #FF8C00; color: white; border-radius: 3px;")
            btn.clicked.connect(lambda _, n=rname, e=exp, c=code, h=handler_name:
                                self._print_label(n, '', e, 0, c,
                                                  label_type='out', handler_name=h))
            self.inventory_table.setCellWidget(idx, 7, btn)
        self.query_count_label.setText(f"共 {len(items)} 筆")
    
    # ================================================================
    # 驗收頁面
    # ================================================================
    def create_qc_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("試劑驗收")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        filter_layout = QHBoxLayout()
        self.qc_show_done = QCheckBox("顯示已完成批次")
        self.qc_show_done.stateChanged.connect(self.refresh_qc_tab)
        filter_layout.addWidget(self.qc_show_done)

        filter_layout.addWidget(QLabel("保管人:"))
        self.qc_keeper_filter = self._get_users_combo()
        self.qc_keeper_filter.setMinimumHeight(34)
        self.qc_keeper_filter.currentIndexChanged.connect(self.refresh_qc_tab)
        filter_layout.addWidget(self.qc_keeper_filter)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.refresh_qc_tab)
        filter_layout.addWidget(refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # 日期分頁容器
        self.qc_date_tabs = QTabWidget()
        layout.addWidget(self.qc_date_tabs)

        self.refresh_qc_tab()
        widget.setLayout(layout)
        return widget

    def refresh_qc_tab(self):
        self.qc_date_tabs.clear()
        try:
            from models import QCBatch
            show_done  = self.qc_show_done.isChecked() if hasattr(self, 'qc_show_done') else False
            keeper_id  = self.qc_keeper_filter.currentData() if hasattr(self, 'qc_keeper_filter') else None

            if show_done:
                grouped = self.qc_service.get_all_batches_by_date()
            else:
                grouped = self.qc_service.get_pending_batches_by_date()

            # 保管人篩選
            if keeper_id:
                filtered = {}
                for date_str, batches in grouped.items():
                    kept = [b for b in batches if b.reagent.keeper_id == keeper_id]
                    if kept:
                        filtered[date_str] = kept
                grouped = filtered

            if not grouped:
                empty = QWidget()
                el = QVBoxLayout(empty)
                el.addWidget(QLabel("目前沒有待驗收項目"))
                self.qc_date_tabs.addTab(empty, "無待驗收")
                return

            for date_str in sorted(grouped.keys(), reverse=True):
                batches = grouped[date_str]
                page    = self._build_qc_date_page(date_str, batches)
                pending = sum(1 for b in batches if not b.is_complete)
                tab_label = f"{date_str}" if pending == 0 else f"{date_str} ({pending}待)"
                self.qc_date_tabs.addTab(page, tab_label)
        except Exception as e:
            print(f"QC tab error: {e}")

    def _build_qc_date_page(self, date_str, batches):
        from models import QCResult
        page = QWidget()
        layout = QVBoxLayout(page)

        # 表格：每批次一行，每個 level 各一欄
        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels([
            "試劑名稱", "LOT號", "Level 數",
            "Level 1", "Level 2", "Level 3", "狀態", "備註"
        ])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(table)

        for row_idx, batch in enumerate(batches):
            table.insertRow(row_idx)
            table.setItem(row_idx, 0, QTableWidgetItem(batch.reagent.reagent_name))
            table.setItem(row_idx, 1, QTableWidgetItem(batch.lot_number))
            table.setItem(row_idx, 2, QTableWidgetItem(str(batch.qc_levels)))

            results = {r.level: r for r in batch.qc_results}
            for lv in range(1, 4):
                if lv > batch.qc_levels:
                    cell = QLineEdit("—")
                    cell.setReadOnly(True)
                    cell.setStyleSheet("background-color:#E0E0E0; border:none;")
                    table.setCellWidget(row_idx, 2 + lv, cell)
                    continue

                r = results.get(lv)
                value_text = (r.value if r and r.value else '')
                edit = QLineEdit(value_text)
                edit.setPlaceholderText("輸入數字或文字")
                edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
                edit.editingFinished.connect(
                    lambda b_id=batch.id, level=lv, w=edit, tbl=table, r_idx=row_idx:
                    self._on_qc_value_changed(b_id, level, w, tbl, r_idx))
                table.setCellWidget(row_idx, 2 + lv, edit)

            # 狀態：下拉選單（待驗收/通過/不通過）
            status_combo = QComboBox()
            status_combo.addItems(["待驗收", "通過", "不通過"])
            overall = self._get_batch_overall_status(batch)
            if overall == 'pass':
                status_combo.setCurrentIndex(1)
            elif overall == 'fail':
                status_combo.setCurrentIndex(2)
            else:
                status_combo.setCurrentIndex(0)
            self._style_qc_status_combo(status_combo)
            status_combo.currentIndexChanged.connect(
                lambda idx, b_id=batch.id, c=status_combo:
                self._on_qc_status_changed(b_id, c))
            table.setCellWidget(row_idx, 6, status_combo)

            # 備註欄
            notes_text = ''
            for r in batch.qc_results:
                if r.notes:
                    notes_text = r.notes
                    break
            note_item = QLineEdit(notes_text)
            note_item.editingFinished.connect(
                lambda b_id=batch.id, w=note_item:
                self._on_qc_notes_changed(b_id, w))
            table.setCellWidget(row_idx, 7, note_item)

        layout.addWidget(QLabel("提示：Level 欄位可輸入任意數值或文字記錄，狀態請用下拉選單選擇，會立即儲存"))
        return page

    def _get_batch_overall_status(self, batch):
        """根據所有 level 的 result 判斷整體狀態"""
        results = batch.qc_results
        if not results:
            return 'pending'
        if any(r.result == 'fail' for r in results):
            return 'fail'
        if all(r.result == 'pass' for r in results):
            return 'pass'
        return 'pending'

    def _style_qc_status_combo(self, combo):
        text = combo.currentText()
        if text == '通過':
            combo.setStyleSheet(
                "QComboBox { background-color:#107C10; color:white; font-weight:bold; }")
        elif text == '不通過':
            combo.setStyleSheet(
                "QComboBox { background-color:#D13438; color:white; font-weight:bold; }")
        else:
            combo.setStyleSheet(
                "QComboBox { background-color:#FFB900; color:black; }")

    def _on_qc_value_changed(self, batch_id, level, widget, table, row_idx):
        """Level 欄位數值變更時儲存（不影響通過/不通過狀態）"""
        value = widget.text()
        ok, msg = self.qc_service.save_value(batch_id, level, value, self.user_id)
        if not ok:
            QMessageBox.critical(self, "錯誤", msg)

    def _on_qc_status_changed(self, batch_id, combo):
        """狀態下拉選單變更：套用到該批次所有 level"""
        text = combo.currentText()
        if text == '通過':
            result = 'pass'
        elif text == '不通過':
            result = 'fail'
        else:
            result = 'pending'

        self._style_qc_status_combo(combo)

        ok, msg = self.qc_service.set_batch_status(batch_id, result, self.user_id)
        if ok:
            self.refresh_qc_tab()
        else:
            QMessageBox.critical(self, "錯誤", msg)

    def _on_qc_notes_changed(self, batch_id, widget):
        """備註變更時，更新該批次所有 level 的備註"""
        try:
            from models import QCResult
            notes = widget.text()
            results = self.session.query(QCResult).filter_by(batch_id=batch_id).all()
            for r in results:
                r.notes = notes
            self.session.commit()
        except Exception as e:
            print(f"備註更新失敗: {e}")

    def _generate_qc_label(self, reagent_name, out_date, in_house_code):
        """驗收標籤：試劑名稱、出庫日期、需驗收大字"""
        try:
            import os
            from utils import LABEL_DIR
            from PIL import Image, ImageDraw, ImageFont
            filename = str(LABEL_DIR / f'qc_label_{in_house_code}_{out_date}.png')

            DPI         = 300
            width       = int(70 * DPI / 25.4)
            height      = int(30 * DPI / 25.4)
            left_margin = int(40 * DPI / 25.4)   # 左邊 20mm 不可印 + 再右移 20mm
            img         = Image.new('RGB', (width, height), 'white')
            draw        = ImageDraw.Draw(img)

            try:
                font_big  = ImageFont.truetype("msjh.ttc", 60)
                font_body = ImageFont.truetype("msjh.ttc", 36)
                font_sm   = ImageFont.truetype("arial.ttf", 28)
            except Exception:
                font_big  = ImageFont.load_default()
                font_body = font_big
                font_sm   = font_big

            # 紅色「需驗收」背景（整列，含左側不可印區一起塗滿較美觀）
            draw.rectangle([0, 0, width, 90], fill='#D13438')
            bbox = draw.textbbox((0, 0), "需  驗  收", font=font_big)
            text_w = bbox[2] - bbox[0]
            usable_width = width - left_margin
            tx = left_margin + (usable_width - text_w) // 2
            draw.text((tx, 10), "需  驗  收", fill='white', font=font_big)

            # 試劑名稱
            draw.text((left_margin + 20, 110), f"試劑：{reagent_name[:18]}", fill='black', font=font_body)
            # 出庫日期
            draw.text((left_margin + 20, 175), f"出庫：{out_date}",          fill='black', font=font_sm)
            # 院內編號
            draw.text((left_margin + 20, 225), f"編號：{in_house_code}",     fill='#555555', font=font_sm)

            draw.line([(0, height - 8), (width, height - 8)], fill='#D13438', width=6)
            img.save(filename)
            return filename
        except Exception as e:
            print(f"驗收標籤生成失敗: {e}")
            return None

    def create_check_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("庫存盤點")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # 篩選列
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("保管人:"))
        self.check_keeper = self._get_users_combo()
        self.check_keeper.setMinimumHeight(34)
        filter_layout.addWidget(self.check_keeper)

        filter_layout.addWidget(QLabel("試劑名稱:"))
        self.check_name_filter = QLineEdit()
        self.check_name_filter.setMinimumHeight(34)
        filter_layout.addWidget(self.check_name_filter)

        refresh_btn = QPushButton("刷新")
        refresh_btn.setMinimumHeight(36)
        refresh_btn.setStyleSheet("background:#0078D4;color:white;border-radius:4px;")
        refresh_btn.clicked.connect(self.refresh_check_table)
        filter_layout.addWidget(refresh_btn)
        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        self.check_count_label = QLabel("共 0 筆")
        layout.addWidget(self.check_count_label)

        self.check_table = QTableWidget()
        self.check_table.setColumnCount(8)
        self.check_table.setHorizontalHeaderLabels([
            "試劑名稱", "LOT號", "單位", "有效期限",
            "系統庫存", "實際庫存", "差異", "reagent_id"
        ])
        self.check_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.check_table)

        # 儲存盤點結果按鈕
        save_layout = QHBoxLayout()
        save_btn = QPushButton("儲存盤點結果")
        save_btn.setMinimumHeight(40)
        save_btn.setStyleSheet("background:#107C10;color:white;border-radius:4px;font-weight:bold;")
        save_btn.clicked.connect(self.save_check_results)
        save_layout.addWidget(save_btn)

        print_btn = QPushButton("列印盤點表")
        print_btn.setMinimumHeight(40)
        print_btn.setStyleSheet("background:#0078D4;color:white;border-radius:4px;font-weight:bold;")
        print_btn.clicked.connect(self.print_check_table)
        save_layout.addWidget(print_btn)

        save_layout.addStretch()
        layout.addLayout(save_layout)

        self.refresh_check_table()
        widget.setLayout(layout)
        return widget
    
    def create_scrap_tab(self):
        """建立報廢標籤"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("報廢管理")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # 表單區
        form_layout = QHBoxLayout()
        
        form_layout.addWidget(QLabel("試劑名稱:"))
        self.scrap_reagent = QComboBox()
        form_layout.addWidget(self.scrap_reagent)
        
        form_layout.addWidget(QLabel("LOT號:"))
        self.scrap_lot = QLineEdit()
        form_layout.addWidget(self.scrap_lot)
        
        form_layout.addWidget(QLabel("數量:"))
        self.scrap_qty = QSpinBox()
        self.scrap_qty.setMinimum(1)
        form_layout.addWidget(self.scrap_qty)
        
        layout.addLayout(form_layout)
        
        # 報廢原因
        reason_layout = QHBoxLayout()
        reason_layout.addWidget(QLabel("報廢原因:"))
        self.scrap_reason = QComboBox()
        self.scrap_reason.addItems(["過期", "變質", "QC失敗", "儀器停用", "其他"])
        reason_layout.addWidget(self.scrap_reason)
        reason_layout.addStretch()
        layout.addLayout(reason_layout)
        
        # 備註
        layout.addWidget(QLabel("備註:"))
        self.scrap_remark = QTextEdit()
        self.scrap_remark.setMaximumHeight(100)
        layout.addWidget(self.scrap_remark)
        
        # 按鈕
        btn_layout = QHBoxLayout()
        
        scrap_btn = QPushButton("執行報廢")
        scrap_btn.setMinimumHeight(45)
        scrap_btn.setStyleSheet("""
            QPushButton {
                background-color: #D13438;
                color: white;
                font-weight: bold;
            }
        """)
        scrap_btn.clicked.connect(self.do_scrap)
        btn_layout.addWidget(scrap_btn)
        
        clear_btn = QPushButton("清除")
        clear_btn.setMinimumHeight(45)
        clear_btn.clicked.connect(self.clear_scrap_form)
        btn_layout.addWidget(clear_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        layout.addStretch()
        
        # 加載試劑清單
        self.refresh_scrap_reagent_combo()
        
        widget.setLayout(layout)
        return widget
    
    def create_traceability_tab(self):
        """建立追溯標籤"""
        widget = QWidget()
        layout = QVBoxLayout()
        
        title = QLabel("試劑追溯")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)
        
        # 查詢條件
        search_layout = QHBoxLayout()
        
        search_layout.addWidget(QLabel("院內編號:"))
        self.trace_code = QLineEdit()
        search_layout.addWidget(self.trace_code)
        
        search_layout.addWidget(QLabel("LOT號:"))
        self.trace_lot = QLineEdit()
        search_layout.addWidget(self.trace_lot)
        
        trace_btn = QPushButton("查詢")
        trace_btn.setMinimumHeight(35)
        trace_btn.clicked.connect(self.do_traceability_query)
        search_layout.addWidget(trace_btn)
        
        layout.addLayout(search_layout)
        
        # 追溯結果
        result_label = QLabel("追溯記錄:")
        result_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(result_label)
        
        self.trace_result = QTextEdit()
        self.trace_result.setReadOnly(True)
        layout.addWidget(self.trace_result)
        
        widget.setLayout(layout)
        return widget
    
    def create_report_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(12)

        title = QLabel("報表列印")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # ── 報表類型切換 ──
        self.report_type_tabs = QTabWidget()
        self.report_type_tabs.addTab(self._build_qc_report_panel(), "驗收完成清單")
        self.report_type_tabs.addTab(self._build_inventory_report_panel(), "庫存列表")
        self.report_type_tabs.addTab(self._build_original_report_panel(), "其他報表")
        layout.addWidget(self.report_type_tabs)

        widget.setLayout(layout)
        return widget

    def _get_users_combo(self):
        """建立使用者下拉，第一項為「不分」"""
        from models import User
        combo = QComboBox()
        combo.addItem("不分", None)
        for u in self.session.query(User).filter_by(is_active=True).all():
            combo.addItem(u.real_name, u.id)
        return combo

    # ── 驗收完成清單面板 ──
    def _build_qc_report_panel(self):
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setSpacing(8)

        # 條件列
        cond = QHBoxLayout()
        cond.addWidget(QLabel("開始日期:"))
        self.qc_rpt_from = QDateEdit()
        self.qc_rpt_from.setCalendarPopup(True)
        self.qc_rpt_from.setDate(QDate.currentDate().addDays(-30))
        self.qc_rpt_from.setMinimumHeight(34)
        cond.addWidget(self.qc_rpt_from)

        cond.addWidget(QLabel("結束日期:"))
        self.qc_rpt_to = QDateEdit()
        self.qc_rpt_to.setCalendarPopup(True)
        self.qc_rpt_to.setDate(QDate.currentDate())
        self.qc_rpt_to.setMinimumHeight(34)
        cond.addWidget(self.qc_rpt_to)

        cond.addWidget(QLabel("保管人:"))
        self.qc_rpt_handler = self._get_users_combo()
        self.qc_rpt_handler.setMinimumHeight(34)
        cond.addWidget(self.qc_rpt_handler)

        preview_btn = QPushButton("預覽")
        preview_btn.setMinimumHeight(36)
        preview_btn.setStyleSheet("background:#0078D4;color:white;border-radius:4px;")
        preview_btn.clicked.connect(self._preview_qc_report)
        cond.addWidget(preview_btn)

        print_btn = QPushButton("列印 PDF")
        print_btn.setMinimumHeight(36)
        print_btn.setStyleSheet("background:#107C10;color:white;border-radius:4px;")
        print_btn.clicked.connect(self._print_qc_report)
        cond.addWidget(print_btn)
        cond.addStretch()
        layout.addLayout(cond)

        self.qc_rpt_count = QLabel("共 0 筆")
        layout.addWidget(self.qc_rpt_count)

        self.qc_rpt_table = QTableWidget()
        self.qc_rpt_table.setColumnCount(8)
        self.qc_rpt_table.setHorizontalHeaderLabels([
            "驗收日期", "試劑名稱", "LOT號", "入庫日期",
            "Level 1", "Level 2", "Level 3", "狀態"
        ])
        self.qc_rpt_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.qc_rpt_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.qc_rpt_table)
        return panel

    def _get_qc_report_data(self):
        """
        回傳已完成的 QCBatch 清單（每批次一筆，包含其所有 QCResult）
        """
        from models import QCBatch, QCResult, ReagentMaster
        from datetime import datetime as dt
        date_from  = self.qc_rpt_from.date().toPyDate()
        date_to    = self.qc_rpt_to.date().toPyDate()
        handler_id = self.qc_rpt_handler.currentData()   # 保管人

        q = (self.session.query(QCBatch)
             .join(ReagentMaster, QCBatch.reagent_id == ReagentMaster.id)
             .filter(QCBatch.is_complete == True))
        if handler_id:
            q = q.filter(ReagentMaster.keeper_id == handler_id)

        batches = q.all()

        # 用每批次最新的 completed_at 作為「驗收日期」並套用日期區間篩選
        result = []
        for b in batches:
            completed_times = [r.completed_at for r in b.qc_results if r.completed_at]
            if not completed_times:
                continue
            latest = max(completed_times)
            if not (dt.combine(date_from, dt.min.time()) <= latest <= dt.combine(date_to, dt.max.time())):
                continue
            result.append((b, latest))

        result.sort(key=lambda x: x[1], reverse=True)
        return result

    def _format_qc_value(self, r):
        """Level 欄位只顯示使用者輸入的數值/文字"""
        if not r:
            return '—'
        return r.value if r.value else ''

    def _preview_qc_report(self):
        self.qc_rpt_table.setRowCount(0)
        rows = self._get_qc_report_data()
        for idx, (batch, latest) in enumerate(rows):
            self.qc_rpt_table.insertRow(idx)
            results = {r.level: r for r in batch.qc_results}

            vals = [
                latest.strftime('%Y-%m-%d %H:%M'),
                batch.reagent.reagent_name, batch.lot_number, batch.stock_in_date,
            ]
            for col, v in enumerate(vals):
                self.qc_rpt_table.setItem(idx, col, QTableWidgetItem(v))

            # Level 1/2/3
            for lv in range(1, 4):
                col = 3 + lv
                if lv > batch.qc_levels:
                    cell = QTableWidgetItem("—")
                    cell.setBackground(QColor('#E0E0E0'))
                else:
                    r = results.get(lv)
                    cell = QTableWidgetItem(self._format_qc_value(r))
                    if r and r.result == 'pass':
                        cell.setBackground(QColor('#107C10'))
                        cell.setForeground(QColor('white'))
                    elif r and r.result == 'fail':
                        cell.setBackground(QColor('#D13438'))
                        cell.setForeground(QColor('white'))
                    else:
                        cell.setBackground(QColor('#FFB900'))
                self.qc_rpt_table.setItem(idx, col, cell)

            # 整體狀態
            overall = self._get_batch_overall_status(batch)
            status_text = {'pass': '通過', 'fail': '不通過', 'pending': '待驗收'}.get(overall, '待驗收')
            s_cell = QTableWidgetItem(status_text)
            if overall == 'pass':
                s_cell.setBackground(QColor('#107C10'))
                s_cell.setForeground(QColor('white'))
            elif overall == 'fail':
                s_cell.setBackground(QColor('#D13438'))
                s_cell.setForeground(QColor('white'))
            else:
                s_cell.setBackground(QColor('#FFB900'))
            self.qc_rpt_table.setItem(idx, 7, s_cell)

        self.qc_rpt_count.setText(f"共 {len(rows)} 筆")

    def _print_qc_report(self):
        rows = self._get_qc_report_data()
        if not rows:
            QMessageBox.warning(self, "提示", "沒有符合條件的資料")
            return
        from_str    = self.qc_rpt_from.date().toString('yyyy-MM-dd')
        to_str      = self.qc_rpt_to.date().toString('yyyy-MM-dd')
        handler_txt = self.qc_rpt_handler.currentText()
        title       = f"驗收完成清單  {from_str} ~ {to_str}  保管人：{handler_txt}"
        headers     = ["驗收日期", "試劑名稱", "LOT號", "入庫日期",
                       "Level 1", "Level 2", "Level 3", "狀態"]
        data_rows   = []
        for batch, latest in rows:
            results = {r.level: r for r in batch.qc_results}
            row = [
                latest.strftime('%Y-%m-%d %H:%M'),
                batch.reagent.reagent_name, batch.lot_number, batch.stock_in_date,
            ]
            for lv in range(1, 4):
                if lv > batch.qc_levels:
                    row.append('—')
                else:
                    r = results.get(lv)
                    row.append(r.value if (r and r.value) else '')
            overall = self._get_batch_overall_status(batch)
            row.append({'pass': '通過', 'fail': '不通過', 'pending': '待驗收'}.get(overall, '待驗收'))
            data_rows.append(row)
        self._export_pdf_table(title, headers, data_rows)

    # ── 庫存列表面板 ──
    def _build_inventory_report_panel(self):
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setSpacing(8)

        cond = QHBoxLayout()
        cond.addWidget(QLabel("保管人:"))
        self.inv_rpt_handler = self._get_users_combo()
        self.inv_rpt_handler.setMinimumHeight(34)
        cond.addWidget(self.inv_rpt_handler)

        cond.addWidget(QLabel("試劑名稱:"))
        self.inv_rpt_name = QLineEdit()
        self.inv_rpt_name.setMinimumHeight(34)
        self.inv_rpt_name.setMinimumWidth(160)
        cond.addWidget(self.inv_rpt_name)

        preview_btn = QPushButton("預覽")
        preview_btn.setMinimumHeight(36)
        preview_btn.setStyleSheet("background:#0078D4;color:white;border-radius:4px;")
        preview_btn.clicked.connect(self._preview_inventory_report)
        cond.addWidget(preview_btn)

        print_btn = QPushButton("列印 PDF")
        print_btn.setMinimumHeight(36)
        print_btn.setStyleSheet("background:#107C10;color:white;border-radius:4px;")
        print_btn.clicked.connect(self._print_inventory_report)
        cond.addWidget(print_btn)
        cond.addStretch()
        layout.addLayout(cond)

        self.inv_rpt_count = QLabel("共 0 筆")
        layout.addWidget(self.inv_rpt_count)

        self.inv_rpt_table = QTableWidget()
        self.inv_rpt_table.setColumnCount(7)
        self.inv_rpt_table.setHorizontalHeaderLabels([
            "試劑名稱", "LOT號", "院內編號", "有效期限",
            "現有庫存", "安全庫存", "狀態"
        ])
        self.inv_rpt_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.inv_rpt_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.inv_rpt_table)
        return panel

    def _get_inventory_report_data(self):
        from models import Inventory, ReagentMaster
        name_filter = self.inv_rpt_name.text().strip()
        handler_id  = self.inv_rpt_handler.currentData()

        q = (self.session.query(Inventory)
             .join(ReagentMaster, Inventory.reagent_id == ReagentMaster.id)
             .filter(Inventory.current_quantity > 0))
        if name_filter:
            q = q.filter(ReagentMaster.reagent_name.contains(name_filter))
        if handler_id:
            q = q.filter(ReagentMaster.keeper_id == handler_id)
        return q.order_by(ReagentMaster.reagent_name).all()

    def _preview_inventory_report(self):
        self.inv_rpt_table.setRowCount(0)
        items = self._get_inventory_report_data()
        for idx, item in enumerate(items):
            self.inv_rpt_table.insertRow(idx)
            status, color = DateHelper.get_expiry_status(item.expiry_date)
            vals = [
                item.reagent.reagent_name, item.lot_number,
                item.in_house_code or '',  item.expiry_date,
                str(item.current_quantity), str(item.reagent.safety_stock), status
            ]
            for col, v in enumerate(vals):
                cell = QTableWidgetItem(v)
                if col == 6:
                    cell.setBackground(QColor(color))
                    cell.setForeground(QColor('white'))
                self.inv_rpt_table.setItem(idx, col, cell)
        self.inv_rpt_count.setText(f"共 {len(items)} 筆")

    def _print_inventory_report(self):
        items = self._get_inventory_report_data()
        if not items:
            QMessageBox.warning(self, "提示", "沒有符合條件的資料")
            return
        handler_txt = self.inv_rpt_handler.currentText()
        title   = f"庫存列表  {datetime.now().strftime('%Y-%m-%d')}  保管人：{handler_txt}"
        headers = ["試劑名稱", "LOT號", "院內編號", "有效期限", "現有", "安全庫存", "狀態"]
        data_rows = []
        for item in items:
            status, _ = DateHelper.get_expiry_status(item.expiry_date)
            data_rows.append([
                item.reagent.reagent_name, item.lot_number,
                item.in_house_code or '', item.expiry_date,
                str(item.current_quantity), str(item.reagent.safety_stock), status
            ])
        self._export_pdf_table(title, headers, data_rows)

    # ── 共用 PDF 列印 ──
    def _export_pdf_table(self, title, headers, data_rows):
        """產生 PDF 並用系統預設程式開啟列印"""
        try:
            import os
            from utils import REPORT_DIR
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.lib import colors as rl_colors
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                            TableStyle, Paragraph, Spacer)
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # 嘗試註冊中文字型（.ttc 需指定 subfontIndex）
            cn_font = 'Helvetica'
            font_candidates = [
                ('MSJH',     'C:/Windows/Fonts/msjh.ttc',     0),
                ('MSJH',     'C:/Windows/Fonts/msjhbd.ttc',   0),
                ('MINGLIU',  'C:/Windows/Fonts/mingliu.ttc',  0),
                ('SIMSUN',   'C:/Windows/Fonts/simsun.ttc',   0),
                ('KAIU',     'C:/Windows/Fonts/kaiu.ttf',     None),
                ('NOTOSANS', 'C:/Windows/Fonts/NotoSansCJK-Regular.ttc', 0),
            ]
            for fname, fpath, sub_idx in font_candidates:
                if os.path.exists(fpath):
                    try:
                        if sub_idx is not None:
                            pdfmetrics.registerFont(TTFont(fname, fpath, subfontIndex=sub_idx))
                        else:
                            pdfmetrics.registerFont(TTFont(fname, fpath))
                        cn_font = fname
                        break
                    except Exception as font_err:
                        print(f"字型載入失敗 {fpath}: {font_err}")
                        continue

            out_path = str(REPORT_DIR / f'report_{datetime.now().strftime("%Y%m%d%H%M%S")}.pdf')

            doc = SimpleDocTemplate(out_path, pagesize=landscape(A4),
                                    leftMargin=10*mm, rightMargin=10*mm,
                                    topMargin=10*mm, bottomMargin=10*mm)
            styles = getSampleStyleSheet()
            story  = []

            # 標題
            from reportlab.platypus import Paragraph
            from reportlab.lib.styles import ParagraphStyle
            title_style = ParagraphStyle('t', fontName=cn_font, fontSize=13,
                                         spaceAfter=6)
            story.append(Paragraph(title, title_style))
            story.append(Spacer(1, 4*mm))

            # 表格
            col_w = (landscape(A4)[0] - 20*mm) / len(headers)
            table_data = [headers] + data_rows
            t = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
            t.setStyle(TableStyle([
                ('FONTNAME',      (0, 0), (-1, -1), cn_font),
                ('FONTSIZE',      (0, 0), (-1, -1), 9),
                ('BACKGROUND',    (0, 0), (-1, 0),  rl_colors.HexColor('#0078D4')),
                ('TEXTCOLOR',     (0, 0), (-1, 0),  rl_colors.white),
                ('FONTSIZE',      (0, 0), (-1, 0),  10),
                ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS',(0, 1), (-1, -1),
                 [rl_colors.white, rl_colors.HexColor('#F0F4F8')]),
                ('GRID',          (0, 0), (-1, -1), 0.5, rl_colors.grey),
                ('ROWHEIGHT',     (0, 0), (-1, -1), 7*mm),
            ]))
            story.append(t)

            # 頁尾
            story.append(Spacer(1, 4*mm))
            foot_style = ParagraphStyle('f', fontName=cn_font, fontSize=8,
                                        textColor=rl_colors.grey)
            story.append(Paragraph(
                f"列印時間：{datetime.now().strftime('%Y-%m-%d %H:%M')}　"
                f"衛生福利部花蓮醫院　共 {len(data_rows)} 筆", foot_style))

            doc.build(story)

            import subprocess
            subprocess.Popen(['start', '', out_path], shell=True)
            QMessageBox.information(self, "完成", f"PDF 已產生並開啟：\n{out_path}")
        except Exception as e:
            QMessageBox.critical(self, "PDF 產生失敗", str(e))

    # ── 其他報表（原本的） ──
    def _build_original_report_panel(self):
        panel = QWidget()
        layout = QVBoxLayout(panel)
        layout.setSpacing(8)

        report_layout = QHBoxLayout()
        report_layout.addWidget(QLabel("報表類型:"))
        self.report_type = QComboBox()
        self.report_type.addItems([
            "庫存報表", "入庫報表", "出庫報表",
            "盤點報表", "報廢報表", "快過期試劑報表"
        ])
        report_layout.addWidget(self.report_type)

        export_btn = QPushButton("匯出 Excel")
        export_btn.setMinimumHeight(40)
        export_btn.clicked.connect(self.export_excel_report)
        report_layout.addWidget(export_btn)

        pdf_btn = QPushButton("匯出 PDF")
        pdf_btn.setMinimumHeight(40)
        pdf_btn.clicked.connect(self.export_pdf_report)
        report_layout.addWidget(pdf_btn)

        report_layout.addStretch()
        layout.addLayout(report_layout)
        layout.addStretch()
        return panel
    
    def create_admin_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        title = QLabel("系統設定")
        title.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        layout.addWidget(title)

        # ── 列印機設定 ──
        printer_label = QLabel("標籤印表機設定")
        printer_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(printer_label)

        printer_layout = QHBoxLayout()
        printer_layout.addWidget(QLabel("印表機名稱:"))
        self.printer_combo = QComboBox()
        self.printer_combo.setMinimumHeight(36)
        self.printer_combo.setMinimumWidth(300)
        self._load_printers()
        printer_layout.addWidget(self.printer_combo)

        refresh_printer_btn = QPushButton("重新整理")
        refresh_printer_btn.setMinimumHeight(36)
        refresh_printer_btn.clicked.connect(self._load_printers)
        printer_layout.addWidget(refresh_printer_btn)

        save_printer_btn = QPushButton("儲存印表機設定")
        save_printer_btn.setMinimumHeight(36)
        save_printer_btn.setStyleSheet("background-color: #0078D4; color: white; border-radius: 4px;")
        save_printer_btn.clicked.connect(self._save_printer)
        printer_layout.addWidget(save_printer_btn)

        printer_layout.addStretch()
        layout.addLayout(printer_layout)

        self.printer_status = QLabel("")
        layout.addWidget(self.printer_status)

        # ── 使用者管理 ──
        user_label = QLabel("使用者管理")
        user_label.setFont(QFont("Arial", 12, QFont.Weight.Bold))
        layout.addWidget(user_label)

        self.user_table = QTableWidget()
        self.user_table.setColumnCount(6)
        self.user_table.setHorizontalHeaderLabels(["ID", "使用者名稱", "姓名", "角色", "部門", "最後登入"])
        self.user_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.user_table.setMaximumHeight(200)
        self.user_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.user_table)

        user_btn_layout = QHBoxLayout()
        add_user_btn = QPushButton("新增使用者")
        add_user_btn.setMinimumHeight(40)
        add_user_btn.clicked.connect(self.add_user_dialog)
        user_btn_layout.addWidget(add_user_btn)
        user_btn_layout.addStretch()
        layout.addLayout(user_btn_layout)

        layout.addStretch()
        self.refresh_user_table()
        widget.setLayout(layout)
        return widget

    def _load_printers(self):
        """從 Windows 讀取所有已安裝印表機"""
        self.printer_combo.clear()
        try:
            import subprocess
            result = subprocess.run(
                ['powershell', '-Command',
                 'Get-Printer | Select-Object -ExpandProperty Name'],
                capture_output=True, text=True, timeout=5
            )
            printers = [p.strip() for p in result.stdout.splitlines() if p.strip()]
            if printers:
                self.printer_combo.addItems(printers)
            else:
                self.printer_combo.addItem("（找不到印表機）")
        except Exception:
            self.printer_combo.addItem("（無法取得印表機列表）")

        # 讀取已儲存的設定
        try:
            from models import SystemSettings
            setting = self.session.query(SystemSettings).filter_by(key='label_printer').first()
            if setting and setting.value:
                idx = self.printer_combo.findText(setting.value)
                if idx >= 0:
                    self.printer_combo.setCurrentIndex(idx)
                self.printer_status.setText(f"目前設定：{setting.value}")
        except Exception:
            pass

    def _save_printer(self):
        """儲存印表機設定"""
        try:
            from models import SystemSettings
            printer_name = self.printer_combo.currentText()
            if '找不到' in printer_name or '無法' in printer_name:
                QMessageBox.warning(self, "錯誤", "請先選擇有效的印表機")
                return
            setting = self.session.query(SystemSettings).filter_by(key='label_printer').first()
            if setting:
                setting.value = printer_name
            else:
                setting = SystemSettings(key='label_printer', value=printer_name)
                self.session.add(setting)
            self.session.commit()
            self.printer_status.setText(f"目前設定：{printer_name}")
            QMessageBox.information(self, "成功", f"已設定印表機：{printer_name}")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    # ============ 事件處理器 ============
    
    def refresh_dashboard(self):
        try:
            from models import Inventory, StockIn, StockOut, ReagentMaster
            from sqlalchemy import func
            from datetime import date as dt_date

            today_start = datetime.combine(datetime.today().date(), datetime.min.time())
            today_end   = datetime.combine(datetime.today().date(), datetime.max.time())

            # 取所有有庫存的項目（後面多處用到）
            from models import Inventory as Inv
            all_inv = self.session.query(Inv).filter(Inv.current_quantity > 0).all()

            # 庫存總數
            total = sum(i.current_quantity for i in all_inv)
            self._stat_cards['total'].setText(str(total))

            # 今日入庫筆數
            in_today = self.session.query(func.count(StockIn.id)).filter(
                StockIn.stock_in_date >= today_start,
                StockIn.stock_in_date <= today_end
            ).scalar() or 0
            self._stat_cards['in_today'].setText(str(in_today))

            # 今日出庫筆數
            out_today = self.session.query(func.count(StockOut.id)).filter(
                StockOut.stock_out_date >= today_start,
                StockOut.stock_out_date <= today_end
            ).scalar() or 0
            self._stat_cards['out_today'].setText(str(out_today))

            # 即將到期（90天內，不含已過期）
            expiring = self.stock_service.get_expiring_items(90)
            self._stat_cards['expiring'].setText(str(len(expiring)))

            # 已過期（現有庫存中已過期的項目）
            expired_count = sum(1 for i in all_inv
                                if DateHelper.get_days_until_expiry(i.expiry_date) < 0)
            self._stat_cards['expired'].setText(str(expired_count))

            # 低庫存
            low = self.stock_service.get_low_stock_items()
            self._stat_cards['low'].setText(str(len(low)))

            # 更新快過期表格（含已過期，顯示全部）
            self.expiry_table.setRowCount(0)
            # 也加入已過期項目
            all_expiry = []
            for i in all_inv:
                days = DateHelper.get_days_until_expiry(i.expiry_date)
                if days < 0:
                    all_expiry.append({
                        'reagent_name': i.reagent.reagent_name,
                        'lot_number': i.lot_number,
                        'expiry_date': i.expiry_date,
                        'remaining_days': days,
                        'quantity': i.current_quantity,
                        'status': 'expired',
                        'color': '#8B0000'
                    })
            combined = sorted(all_expiry + expiring, key=lambda x: x['remaining_days'])
            for idx, item in enumerate(combined):
                self.expiry_table.insertRow(idx)
                days = item['remaining_days']
                # 顏色分層：已過期/7天/14天/30天/90天
                if days < 0:
                    color = '#8B0000'
                elif days <= 7:
                    color = '#CC0000'
                elif days <= 14:
                    color = '#D13438'
                elif days <= 30:
                    color = '#FF6600'
                else:
                    color = '#FFB900'
                self.expiry_table.setItem(idx, 0, QTableWidgetItem(item['reagent_name']))
                self.expiry_table.setItem(idx, 1, QTableWidgetItem(item['lot_number']))
                self.expiry_table.setItem(idx, 2, QTableWidgetItem(item['expiry_date']))
                days_item = QTableWidgetItem(str(days) if days >= 0 else f"已過期{abs(days)}天")
                days_item.setBackground(QColor(color))
                days_item.setForeground(QColor('white'))
                self.expiry_table.setItem(idx, 3, days_item)
                self.expiry_table.setItem(idx, 4, QTableWidgetItem(str(item['quantity'])))

            # 更新低庫存表格（全部顯示，不限筆數）
            self.low_stock_table.setRowCount(0)
            for idx, item in enumerate(low):
                self.low_stock_table.insertRow(idx)
                self.low_stock_table.setItem(idx, 0, QTableWidgetItem(item['reagent_name']))
                self.low_stock_table.setItem(idx, 1, QTableWidgetItem(str(item['current'])))
                self.low_stock_table.setItem(idx, 2, QTableWidgetItem(str(item['safety'])))
                s = QTableWidgetItem(str(item['shortage']))
                s.setBackground(QColor('#D13438'))
                s.setForeground(QColor('white'))
                self.low_stock_table.setItem(idx, 3, s)

        except Exception as e:
            print(f"Dashboard error: {e}")
    
    def refresh_reagent_table(self):
        self.reagent_table.setRowCount(0)
        reagents = self.reagent_service.get_all_reagents()
        for idx, r in enumerate(reagents):
            self.reagent_table.insertRow(idx)
            self.reagent_table.setItem(idx, 0,  QTableWidgetItem(str(r.id)))
            self.reagent_table.setItem(idx, 1,  QTableWidgetItem(r.reagent_code))
            self.reagent_table.setItem(idx, 2,  QTableWidgetItem(r.reagent_name))
            self.reagent_table.setItem(idx, 3,  QTableWidgetItem(r.brand or ''))
            self.reagent_table.setItem(idx, 4,  QTableWidgetItem(r.supplier or ''))
            self.reagent_table.setItem(idx, 5,  QTableWidgetItem(r.specification or ''))
            self.reagent_table.setItem(idx, 6,  QTableWidgetItem(r.unit or ''))
            self.reagent_table.setItem(idx, 7,  QTableWidgetItem(str(r.safety_stock)))
            self.reagent_table.setItem(idx, 8,  QTableWidgetItem(str(r.lot_start  if r.lot_start  is not None else 0)))
            self.reagent_table.setItem(idx, 9,  QTableWidgetItem(str(r.lot_length if r.lot_length is not None else 10)))
            self.reagent_table.setItem(idx, 10, QTableWidgetItem(str(r.exp_start  if r.exp_start  is not None else 10)))
            self.reagent_table.setItem(idx, 11, QTableWidgetItem(str(r.exp_length if r.exp_length is not None else 6)))
            self.reagent_table.setItem(idx, 12, QTableWidgetItem(r.exp_format or 'YYYYMMDD'))
            keeper_name = r.keeper.real_name if r.keeper else ''
            self.reagent_table.setItem(idx, 13, QTableWidgetItem(keeper_name))

    def download_reagent_template(self):
        """下載試劑匯入 Excel 範本"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "儲存範本", "試劑匯入範本.xlsx", "Excel Files (*.xlsx)"
            )
            if not file_path:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "試劑清單"

            headers = [
                "條碼前16碼*", "試劑名稱*", "廠牌", "供應商",
                "規格", "單位", "安全庫存",
                "批號起始位置", "批號長度", "效期起始位置", "效期長度", "效期格式",
                "需要驗收(是/否)", "品管Level數(1-3)", "每盒數量", "保管人(輸入姓名)"
            ]

            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # 範例資料
            example = [
                "1234567890123456", "葡萄糖試劑", "華東", "華東醫材",
                "100ml/瓶", "瓶", 10,
                0, 10, 10, 6, "YYYYMMDD",
                "是", 2, 1, ""
            ]
            for col, v in enumerate(example, start=1):
                ws.cell(row=2, column=col, value=v)

            # 欄寬
            widths = [18, 20, 12, 16, 16, 8, 10, 10, 10, 10, 10, 12, 14, 14, 10, 14]
            for col, w in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

            # 說明頁
            ws2 = wb.create_sheet("說明")
            notes = [
                "欄位說明：",
                "1. 條碼前16碼*：必填，需唯一，用於入庫掃描辨識試劑",
                "2. 試劑名稱*：必填",
                "3. 批號起始位置/長度：條碼中批號的位置（從0開始算）",
                "4. 效期起始位置/長度/格式：條碼中效期的位置與格式",
                "   效期格式可選: YYYYMMDD, YYMMDD, YYYY-MM-DD, YY/MM/DD, MMDDYYYY, DDMMYYYY",
                "5. 需要驗收：填「是」或「否」",
                "6. 品管Level數：1~3，僅在需要驗收=是時生效",
                "7. 每盒數量：掃描一次入庫代表幾個，預設為1",
                "8. 保管人：填寫系統中已存在的使用者姓名，留空則不指定",
                "",
                "* 為必填欄位，其他欄位可留空",
                "上傳時，若條碼前16碼已存在，會更新該筆試劑資料；若不存在則新增"
            ]
            for i, line in enumerate(notes, start=1):
                ws2.cell(row=i, column=1, value=line)
            ws2.column_dimensions['A'].width = 70

            wb.save(file_path)
            QMessageBox.information(self, "完成", f"範本已儲存：\n{file_path}")

            import subprocess
            subprocess.Popen(['start', '', file_path], shell=True)

        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def import_reagent_excel(self):
        """從 Excel 匯入試劑資料"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self, "選擇試劑匯入檔案", "", "Excel Files (*.xlsx *.xls)"
            )
            if not file_path:
                return

            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb["試劑清單"] if "試劑清單" in wb.sheetnames else wb.active

            from models import User as U

            success_count = 0
            update_count   = 0
            error_rows     = []

            for row_idx in range(2, ws.max_row + 1):
                row = [ws.cell(row=row_idx, column=c).value for c in range(1, 17)]
                if all(v is None or str(v).strip() == '' for v in row):
                    continue  # 跳過空白行

                (code, name, brand, supplier, spec, unit, safety,
                 lot_start, lot_len, exp_start, exp_len, exp_fmt,
                 need_qc_text, qc_levels, units_per_box, keeper_name) = (
                    row + [None] * (16 - len(row))
                )[:16]

                code = str(code).strip() if code else ''
                name = str(name).strip() if name else ''

                if not code or not name:
                    error_rows.append(f"第 {row_idx} 行：條碼前16碼或試劑名稱為空，已跳過")
                    continue

                def to_int(v, default):
                    try:
                        if v is None or str(v).strip() == '':
                            return default
                        return int(float(v))
                    except Exception:
                        return default

                need_qc = str(need_qc_text).strip() in ('是', 'Y', 'y', 'TRUE', 'true', '1') if need_qc_text else False

                # 保管人
                keeper_id = None
                if keeper_name and str(keeper_name).strip():
                    u = self.session.query(U).filter_by(
                        real_name=str(keeper_name).strip(), is_active=True).first()
                    if u:
                        keeper_id = u.id

                data = dict(
                    reagent_code=code,
                    reagent_name=name,
                    brand=str(brand) if brand else '',
                    supplier=str(supplier) if supplier else '',
                    specification=str(spec) if spec else '',
                    unit=str(unit) if unit else '',
                    safety_stock=to_int(safety, 10),
                    lot_start=to_int(lot_start, 0),
                    lot_length=to_int(lot_len, 10),
                    exp_start=to_int(exp_start, 10),
                    exp_length=to_int(exp_len, 6),
                    exp_format=str(exp_fmt).strip() if exp_fmt else 'YYYYMMDD',
                    need_qc=need_qc,
                    qc_levels=to_int(qc_levels, 1),
                    units_per_box=to_int(units_per_box, 1),
                    keeper_id=keeper_id
                )

                existing = self.reagent_service.get_reagent_by_code(code)
                if existing:
                    ok, msg = self.reagent_service.update_reagent(existing.id, **data)
                    if ok:
                        update_count += 1
                    else:
                        error_rows.append(f"第 {row_idx} 行：{msg}")
                else:
                    ok, msg = self.reagent_service.create_reagent(**data)
                    if ok:
                        success_count += 1
                    else:
                        error_rows.append(f"第 {row_idx} 行：{msg}")

            self.refresh_reagent_table()

            result_msg = f"匯入完成！\n新增：{success_count} 筆\n更新：{update_count} 筆"
            if error_rows:
                result_msg += "\n\n錯誤訊息：\n" + "\n".join(error_rows[:10])
                if len(error_rows) > 10:
                    result_msg += f"\n...等共 {len(error_rows)} 個錯誤"
            QMessageBox.information(self, "匯入結果", result_msg)

        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    def refresh_scrap_reagent_combo(self):
        """刷新報廢試劑下拉清單"""
        self.scrap_reagent.clear()
        reagents = self.reagent_service.get_all_reagents()
        for reagent in reagents:
            self.scrap_reagent.addItem(reagent.reagent_name, reagent.id)
    
    def refresh_check_table(self):
        self.check_table.setRowCount(0)
        from models import Inventory, ReagentMaster, StockIn
        keeper_id   = self.check_keeper.currentData() if hasattr(self, 'check_keeper') else None
        name_filter = self.check_name_filter.text().strip() if hasattr(self, 'check_name_filter') else ''

        # 取所有試劑
        rq = self.session.query(ReagentMaster).filter_by(is_active=True)
        if keeper_id:
            rq = rq.filter(ReagentMaster.keeper_id == keeper_id)
        if name_filter:
            rq = rq.filter(ReagentMaster.reagent_name.contains(name_filter))
        reagents = rq.order_by(ReagentMaster.reagent_name).all()

        # 取所有庫存（依 reagent+lot+expiry 分組加總，去除空白避免無法合併）
        all_inv = self.session.query(Inventory).all()
        grouped = {}  # (reagent_id, lot_number, expiry_date) -> total_qty
        for inv in all_inv:
            lot = (inv.lot_number or '').strip()
            exp = (inv.expiry_date or '').strip()
            key = (inv.reagent_id, lot, exp)
            grouped[key] = grouped.get(key, 0) + inv.current_quantity

        # 組合顯示資料：每個試劑 × 每個有庫存的 LOT（庫存為 0 的批號不顯示）
        rows_data = []
        reagent_ids = {r.id for r in reagents}
        for (reagent_id, lot_number, expiry_date), total_qty in grouped.items():
            if reagent_id not in reagent_ids:
                continue
            if total_qty == 0:
                continue  # 數量為 0 的批號不列出
            rows_data.append((reagent_id, lot_number, expiry_date, total_qty))

        # 每個試劑至少要保留一列；若該試劑所有批號都被過濾掉，補一列數量為 0
        reagent_has_row = {rid for (rid, _, _, _) in rows_data}
        for r in reagents:
            if r.id not in reagent_has_row:
                rows_data.append((r.id, '', '', 0))

        # 排序：依試劑名稱、效期
        reagent_map = {r.id: r for r in reagents}
        rows_data.sort(key=lambda x: (
            reagent_map[x[0]].reagent_name if x[0] in reagent_map else '',
            x[2] or ''
        ))

        self.check_table.setRowCount(0)
        for idx, (reagent_id, lot_number, expiry_date, total_qty) in enumerate(rows_data):
            reagent = reagent_map.get(reagent_id)
            if not reagent:
                continue
            self.check_table.insertRow(idx)
            self.check_table.setItem(idx, 0, QTableWidgetItem(reagent.reagent_name))
            self.check_table.setItem(idx, 1, QTableWidgetItem(lot_number))
            self.check_table.setItem(idx, 2, QTableWidgetItem(reagent.unit or ''))
            self.check_table.setItem(idx, 3, QTableWidgetItem(expiry_date or ''))
            self.check_table.setItem(idx, 4, QTableWidgetItem(str(total_qty)))

            # 隱藏資料：用於儲存比對
            id_item = QTableWidgetItem(str(reagent_id))
            self.check_table.setItem(idx, 7, id_item)  # 隱藏欄存 reagent_id

            # 實際庫存可編輯
            spin = QSpinBox()
            spin.setRange(0, 99999)
            spin.setValue(total_qty)
            self.check_table.setCellWidget(idx, 5, spin)

            diff = QTableWidgetItem("0")
            self.check_table.setItem(idx, 6, diff)

            def make_diff_updater(row_i, sys_qty, spin_w):
                def update():
                    d = spin_w.value() - sys_qty
                    di = self.check_table.item(row_i, 6)
                    if di:
                        di.setText(str(d))
                        di.setBackground(QColor('#D13438' if d < 0 else ('#107C10' if d > 0 else '#FFFFFF')))
                        di.setForeground(QColor('white' if d != 0 else 'black'))
                return update
            spin.valueChanged.connect(make_diff_updater(idx, total_qty, spin))

        # 隱藏第 7 欄（reagent_id）
        self.check_table.setColumnHidden(7, True)

        if hasattr(self, 'check_count_label'):
            self.check_count_label.setText(f"共 {len(rows_data)} 筆")

    def save_check_results(self):
        """儲存盤點結果：差異會平均分配到該批號底下各個 Inventory 記錄"""
        try:
            from models import Inventory, InventoryCheck
            count = 0
            for row in range(self.check_table.rowCount()):
                spin = self.check_table.cellWidget(row, 5)
                if not spin:
                    continue
                actual = spin.value()

                reagent_id_item = self.check_table.item(row, 7)
                lot_item        = self.check_table.item(row, 1)
                exp_item        = self.check_table.item(row, 3)
                if not reagent_id_item or not lot_item:
                    continue
                reagent_id  = int(reagent_id_item.text())
                lot_number  = lot_item.text().strip()
                expiry_date = (exp_item.text().strip() if exp_item else '')

                if not lot_number:
                    continue  # 無庫存紀錄的試劑，不寫入

                # 找出該批號所有 Inventory 記錄（strip 比對避免空白差異）
                candidates = self.session.query(Inventory).filter_by(
                    reagent_id=reagent_id
                ).all()
                invs = [i for i in candidates
                        if (i.lot_number or '').strip() == lot_number
                        and (i.expiry_date or '').strip() == expiry_date]
                if not invs:
                    continue

                system_qty = sum(i.current_quantity for i in invs)
                diff = actual - system_qty
                if diff == 0:
                    continue

                # 寫入盤點紀錄
                record = InventoryCheck(
                    reagent_id=reagent_id,
                    lot_number=lot_number,
                    system_quantity=system_qty,
                    actual_quantity=actual,
                    difference=diff
                )
                self.session.add(record)

                # 調整庫存：差異加到第一筆，若導致負數則往後找補
                remaining = diff
                for inv in invs:
                    if remaining == 0:
                        break
                    new_qty = inv.current_quantity + remaining
                    if new_qty < 0:
                        remaining = new_qty
                        inv.current_quantity = 0
                    else:
                        inv.current_quantity = new_qty
                        remaining = 0
                # 若仍有剩餘負數（全部已是0），忽略（極端情況）

                count += 1

            self.session.commit()
            QMessageBox.information(self, "完成", f"已儲存 {count} 筆有差異的盤點結果")
            self.refresh_check_table()
            self.refresh_dashboard()
        except Exception as e:
            self.session.rollback()
            QMessageBox.critical(self, "錯誤", str(e))

    def print_check_table(self):
        """列印盤點表 PDF"""
        rows = []
        for row in range(self.check_table.rowCount()):
            name  = self.check_table.item(row, 0).text()
            lot   = self.check_table.item(row, 1).text()
            unit  = self.check_table.item(row, 2).text()
            exp   = self.check_table.item(row, 3).text()
            sysq  = self.check_table.item(row, 4).text()
            spin  = self.check_table.cellWidget(row, 5)
            actual = str(spin.value()) if spin else sysq
            diff  = self.check_table.item(row, 6).text()
            rows.append([name, lot, unit, exp, sysq, actual, diff])

        if not rows:
            QMessageBox.warning(self, "提示", "沒有資料可列印")
            return

        keeper_txt = self.check_keeper.currentText()
        title = f"庫存盤點表  {datetime.now().strftime('%Y-%m-%d')}  保管人：{keeper_txt}"
        headers = ["試劑名稱", "LOT號", "單位", "有效期限", "系統庫存", "實際庫存", "差異"]
        self._export_pdf_table(title, headers, rows)
    
    def refresh_user_table(self):
        """刷新使用者表格"""
        self.user_table.setRowCount(0)
        users = self.user_service.get_all_users()
        
        for idx, user in enumerate(users):
            self.user_table.insertRow(idx)
            self.user_table.setItem(idx, 0, QTableWidgetItem(str(user.id)))
            self.user_table.setItem(idx, 1, QTableWidgetItem(user.username))
            self.user_table.setItem(idx, 2, QTableWidgetItem(user.real_name))
            self.user_table.setItem(idx, 3, QTableWidgetItem(user.role))
            self.user_table.setItem(idx, 4, QTableWidgetItem(user.department or ''))
            self.user_table.setItem(idx, 5, QTableWidgetItem(
                user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else '-'
            ))
    
    def refresh_audit_log_table(self):
        """刷新審計日誌表格"""
        # 由於篇幅限制，這裡簡化實現
        pass
    
    def _reagent_dialog(self, title, reagent=None):
        """共用的新增/編輯試劑對話框"""
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setMinimumWidth(520)
        layout = QVBoxLayout()
        layout.setSpacing(8)

        def row(label, widget):
            h = QHBoxLayout()
            lbl = QLabel(label)
            lbl.setMinimumWidth(110)
            h.addWidget(lbl)
            h.addWidget(widget)
            layout.addLayout(h)

        f_code    = QLineEdit(reagent.reagent_code    if reagent else '')
        f_name    = QLineEdit(reagent.reagent_name    if reagent else '')
        f_brand   = QLineEdit(reagent.brand           if reagent else '')
        f_supplier= QLineEdit(reagent.supplier        if reagent else '')
        f_spec    = QLineEdit(reagent.specification   if reagent else '')
        f_unit    = QLineEdit(reagent.unit            if reagent else '')
        f_safety  = QSpinBox(); f_safety.setRange(0, 99999)
        f_safety.setValue(reagent.safety_stock if reagent else 10)
        f_storage = QLineEdit(reagent.storage_condition if reagent else '')
        f_remark  = QLineEdit(reagent.remark          if reagent else '')

        # 條碼解析設定
        sep = QLabel("── 條碼解析設定（掃描條碼自動取批號/效期）──")
        sep.setStyleSheet("color: #0078D4; font-weight: bold;")

        f_lot_start  = QSpinBox(); f_lot_start.setRange(0, 99)
        f_lot_start.setValue(reagent.lot_start  if reagent and reagent.lot_start  is not None else 0)
        f_lot_len    = QSpinBox(); f_lot_len.setRange(1, 50)
        f_lot_len.setValue(reagent.lot_length   if reagent and reagent.lot_length is not None else 10)
        f_exp_start  = QSpinBox(); f_exp_start.setRange(0, 99)
        f_exp_start.setValue(reagent.exp_start  if reagent and reagent.exp_start  is not None else 10)
        f_exp_len    = QSpinBox(); f_exp_len.setRange(1, 20)
        f_exp_len.setValue(reagent.exp_length   if reagent and reagent.exp_length is not None else 6)
        f_exp_fmt    = QComboBox()
        f_exp_fmt.addItems(['YYYYMMDD', 'YYMMDD', 'YYYY-MM-DD', 'YY/MM/DD', 'MMDDYYYY', 'DDMMYYYY'])
        if reagent and reagent.exp_format:
            idx = f_exp_fmt.findText(reagent.exp_format)
            if idx >= 0:
                f_exp_fmt.setCurrentIndex(idx)

        row("條碼前16碼:",  f_code)
        row("試劑名稱:",     f_name)
        row("廠牌:",         f_brand)
        row("供應商:",       f_supplier)
        row("規格:",         f_spec)
        row("單位:",         f_unit)
        row("安全庫存:",     f_safety)
        row("儲存條件:",     f_storage)
        row("備註:",         f_remark)
        layout.addWidget(sep)

        hint = QLabel("例：條碼 'ABC123LOT456EXP20261231'\n批號起始=6, 批號長度=6, 效期起始=15, 效期長度=8, 效期格式=YYYYMMDD")
        hint.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(hint)

        row("批號起始位置:", f_lot_start)
        row("批號長度:",     f_lot_len)
        row("效期起始位置:", f_exp_start)
        row("效期長度:",     f_exp_len)
        row("效期格式:",     f_exp_fmt)

        # 驗收設定
        sep2 = QLabel("── 驗收設定 ──")
        sep2.setStyleSheet("color: #D13438; font-weight: bold;")
        layout.addWidget(sep2)

        qc_layout = QHBoxLayout()
        f_need_qc = QCheckBox("需要驗收")
        f_need_qc.setChecked(reagent.need_qc if reagent else False)
        qc_layout.addWidget(f_need_qc)

        qc_layout.addWidget(QLabel("品管 Level 數:"))
        f_qc_levels = QComboBox()
        f_qc_levels.addItems(["1", "2", "3"])
        f_qc_levels.setCurrentIndex((reagent.qc_levels - 1) if reagent and reagent.qc_levels else 0)
        f_qc_levels.setEnabled(f_need_qc.isChecked())
        f_need_qc.stateChanged.connect(lambda s: f_qc_levels.setEnabled(bool(s)))
        qc_layout.addWidget(f_qc_levels)
        qc_layout.addStretch()
        layout.addLayout(qc_layout)

        # 保管人設定
        sep3 = QLabel("── 保管人 ──")
        sep3.setStyleSheet("color: #107C10; font-weight: bold;")
        layout.addWidget(sep3)

        keeper_layout = QHBoxLayout()
        keeper_layout.addWidget(QLabel("保管人:"))
        f_keeper = QComboBox()
        f_keeper.addItem("未指定", None)
        from models import User as U
        for u in self.session.query(U).filter_by(is_active=True).all():
            f_keeper.addItem(u.real_name, u.id)
        # 預設選目前保管人
        if reagent and reagent.keeper_id:
            idx = f_keeper.findData(reagent.keeper_id)
            if idx >= 0:
                f_keeper.setCurrentIndex(idx)
        keeper_layout.addWidget(f_keeper)
        keeper_layout.addStretch()
        layout.addLayout(keeper_layout)

        # 包裝數量設定
        sep4 = QLabel("── 包裝設定 ──")
        sep4.setStyleSheet("color: #FF8C00; font-weight: bold;")
        layout.addWidget(sep4)

        box_layout = QHBoxLayout()
        box_layout.addWidget(QLabel("每盒數量:"))
        f_units_per_box = QSpinBox()
        f_units_per_box.setRange(1, 9999)
        f_units_per_box.setValue(reagent.units_per_box if reagent and reagent.units_per_box else 1)
        box_layout.addWidget(f_units_per_box)
        hint2 = QLabel("（掃描一次條碼＝一盒，會自動產生對應數量的獨立院內編號與標籤）")
        hint2.setStyleSheet("color: #666; font-size: 10px;")
        box_layout.addWidget(hint2)
        box_layout.addStretch()
        layout.addLayout(box_layout)

        # 按鈕
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("儲存")
        save_btn.setMinimumHeight(40)
        save_btn.setStyleSheet("background-color: #0078D4; color: white; font-weight: bold; border-radius: 4px;")
        cancel_btn = QPushButton("取消")
        cancel_btn.setMinimumHeight(40)
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        dialog.setLayout(layout)

        result = {'saved': False}

        def save():
            code = f_code.text().strip()
            name = f_name.text().strip()
            if not code or not name:
                QMessageBox.warning(dialog, "錯誤", "條碼前16碼和試劑名稱為必填")
                return
            data = dict(
                reagent_code=code, reagent_name=name,
                brand=f_brand.text(), supplier=f_supplier.text(),
                specification=f_spec.text(), unit=f_unit.text(),
                safety_stock=f_safety.value(), storage_condition=f_storage.text(),
                remark=f_remark.text(),
                lot_start=f_lot_start.value(), lot_length=f_lot_len.value(),
                exp_start=f_exp_start.value(), exp_length=f_exp_len.value(),
                exp_format=f_exp_fmt.currentText(),
                need_qc=f_need_qc.isChecked(),
                qc_levels=int(f_qc_levels.currentText()),
                keeper_id=f_keeper.currentData(),
                units_per_box=f_units_per_box.value()
            )
            if reagent:
                ok, msg = self.reagent_service.update_reagent(reagent.id, **data)
            else:
                ok, msg = self.reagent_service.create_reagent(**data)
            if ok:
                result['saved'] = True
                self.refresh_reagent_table()
                dialog.accept()
            else:
                QMessageBox.critical(dialog, "錯誤", msg)

        save_btn.clicked.connect(save)
        dialog.exec()
        return result['saved']

    def add_reagent_dialog(self):
        self._reagent_dialog("新增試劑")

    def edit_reagent_dialog(self):
        rows = self.reagent_table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "提示", "請先選取要編輯的試劑")
            return
        reagent_id = int(self.reagent_table.item(rows[0].row(), 0).text())
        reagent = self.reagent_service.get_reagent_by_id(reagent_id)
        if reagent:
            self._reagent_dialog("編輯試劑", reagent)

    def delete_reagent(self):
        rows = self.reagent_table.selectionModel().selectedRows()
        if not rows:
            QMessageBox.warning(self, "提示", "請先選取要刪除的試劑")
            return
        reagent_id = int(self.reagent_table.item(rows[0].row(), 0).text())
        name = self.reagent_table.item(rows[0].row(), 2).text()
        reply = QMessageBox.question(self, "確認", f"確定刪除「{name}」？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            ok, msg = self.reagent_service.delete_reagent(reagent_id)
            QMessageBox.information(self, "結果", msg)
            self.refresh_reagent_table()

    def open_reagent_search_dialog(self):
        """開啟試劑搜尋視窗，選擇後帶入入庫條碼欄位"""
        from models import ReagentMaster, Inventory
        from sqlalchemy import func

        dialog = QDialog(self)
        dialog.setWindowTitle("搜尋試劑")
        dialog.resize(900, 600)
        layout = QVBoxLayout(dialog)

        # ── 搜尋條件列 ──
        cond_layout = QHBoxLayout()

        cond_layout.addWidget(QLabel("試劑名稱:"))
        f_name = QLineEdit()
        f_name.setMinimumHeight(34)
        f_name.setMinimumWidth(150)
        cond_layout.addWidget(f_name)

        cond_layout.addWidget(QLabel("廠牌:"))
        f_brand = QLineEdit()
        f_brand.setMinimumHeight(34)
        f_brand.setMinimumWidth(100)
        cond_layout.addWidget(f_brand)

        cond_layout.addWidget(QLabel("條碼:"))
        f_code = QLineEdit()
        f_code.setMinimumHeight(34)
        f_code.setMinimumWidth(140)
        cond_layout.addWidget(f_code)

        cond_layout.addWidget(QLabel("供應商:"))
        f_supplier = QLineEdit()
        f_supplier.setMinimumHeight(34)
        f_supplier.setMinimumWidth(100)
        cond_layout.addWidget(f_supplier)

        search_btn2 = QPushButton("搜尋")
        search_btn2.setMinimumHeight(34)
        search_btn2.setStyleSheet("background:#0078D4; color:white; border-radius:4px;")
        cond_layout.addWidget(search_btn2)

        clear_btn2 = QPushButton("清除")
        clear_btn2.setMinimumHeight(34)
        cond_layout.addWidget(clear_btn2)

        cond_layout.addStretch()
        layout.addLayout(cond_layout)

        count_lbl = QLabel("共 0 筆")
        layout.addWidget(count_lbl)

        # ── 搜尋結果表格 ──
        table = QTableWidget()
        table.setColumnCount(7)
        table.setHorizontalHeaderLabels([
            "條碼前16碼", "試劑名稱", "廠牌", "供應商", "規格", "單位", "現有庫存"
        ])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(table)

        hint = QLabel("雙擊或選取後按「確認選擇」帶入條碼")
        hint.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(hint)

        # ── 底部按鈕 ──
        bottom = QHBoxLayout()
        confirm_btn = QPushButton("確認選擇")
        confirm_btn.setMinimumHeight(40)
        confirm_btn.setStyleSheet("background:#107C10; color:white; font-weight:bold; border-radius:4px;")
        cancel_btn2 = QPushButton("取消")
        cancel_btn2.setMinimumHeight(40)
        cancel_btn2.clicked.connect(dialog.reject)
        bottom.addWidget(confirm_btn)
        bottom.addWidget(cancel_btn2)
        bottom.addStretch()
        layout.addLayout(bottom)

        # ── 取得庫存加總 (reagent_id -> total) ──
        inv_totals = {}
        for inv in self.session.query(Inventory).all():
            inv_totals[inv.reagent_id] = inv_totals.get(inv.reagent_id, 0) + inv.current_quantity

        def do_search():
            table.setRowCount(0)
            q = self.session.query(ReagentMaster).filter_by(is_active=True)
            if f_name.text().strip():
                q = q.filter(ReagentMaster.reagent_name.contains(f_name.text().strip()))
            if f_brand.text().strip():
                q = q.filter(ReagentMaster.brand.contains(f_brand.text().strip()))
            if f_code.text().strip():
                q = q.filter(ReagentMaster.reagent_code.contains(f_code.text().strip()))
            if f_supplier.text().strip():
                q = q.filter(ReagentMaster.supplier.contains(f_supplier.text().strip()))
            reagents = q.order_by(ReagentMaster.reagent_name).all()
            for idx, r in enumerate(reagents):
                table.insertRow(idx)
                total = inv_totals.get(r.id, 0)
                for col, v in enumerate([
                    r.reagent_code, r.reagent_name, r.brand or '',
                    r.supplier or '', r.specification or '',
                    r.unit or '', str(total)
                ]):
                    item = QTableWidgetItem(v)
                    if col == 6 and total <= r.safety_stock:
                        item.setBackground(QColor('#FFB900'))
                    table.setItem(idx, col, item)
            count_lbl.setText(f"共 {len(reagents)} 筆")

        def on_confirm():
            rows = table.selectionModel().selectedRows()
            if not rows:
                QMessageBox.warning(dialog, "提示", "請先選取一個試劑")
                return
            code = table.item(rows[0].row(), 0).text()
            self.barcode_input.setText(code)
            self.parse_scanned_barcode()
            # 清空批號和效期讓人員手動輸入
            self.stock_in_lot.clear()
            self.stock_in_expiry.clear()
            self.stock_in_lot.setFocus()
            dialog.accept()

        search_btn2.clicked.connect(do_search)
        confirm_btn.clicked.connect(on_confirm)
        table.doubleClicked.connect(on_confirm)
        clear_btn2.clicked.connect(lambda: [
            f_name.clear(), f_brand.clear(), f_code.clear(), f_supplier.clear()
        ])
        f_name.returnPressed.connect(do_search)

        # 預設顯示所有試劑
        do_search()
        dialog.exec()

    def parse_scanned_barcode(self):
        """掃描條碼後自動識別試劑、批號、效期"""
        raw = self.barcode_input.text().strip()
        if not raw:
            return

        # 取前16碼比對試劑
        prefix = raw[:16]
        from models import ReagentMaster
        reagent = self.session.query(ReagentMaster).filter(
            ReagentMaster.reagent_code == prefix,
            ReagentMaster.is_active == True
        ).first()

        if not reagent:
            # 嘗試用更短的前綴比對（依序從16碼縮短）
            for length in range(15, 3, -1):
                p = raw[:length]
                reagent = self.session.query(ReagentMaster).filter(
                    ReagentMaster.reagent_code == p,
                    ReagentMaster.is_active == True
                ).first()
                if reagent:
                    break

        if not reagent:
            QMessageBox.warning(self, "未找到試劑",
                f"條碼前16碼「{prefix}」無對應試劑\n請先在試劑設定中新增此試劑")
            return

        self.stock_in_reagent_id = reagent.id
        self.stock_in_reagent_name.setText(reagent.reagent_name)

        # 解析批號
        lot_start  = reagent.lot_start  or 0
        lot_length = reagent.lot_length or 10
        lot = raw[lot_start: lot_start + lot_length].strip()
        self.stock_in_lot.setText(lot)

        # 解析效期
        exp_start  = reagent.exp_start  or 10
        exp_length = reagent.exp_length or 6
        exp_raw = raw[exp_start: exp_start + exp_length].strip()
        exp_date = self._parse_exp_date(exp_raw, reagent.exp_format or 'YYYYMMDD')
        self.stock_in_expiry.setText(exp_date)

        # 顯示每盒數量提示
        units = reagent.units_per_box or 1
        if units > 1:
            self.stock_in_box_hint.setText(f"× 每盒 {units} 個 = {self.stock_in_qty.value() * units} 個")
        else:
            self.stock_in_box_hint.setText("")

        # 即時更新提示
        try:
            self.stock_in_qty.valueChanged.disconnect()
        except Exception:
            pass
        self.stock_in_qty.lineEdit().returnPressed.connect(self.do_stock_in)

        def update_hint(v):
            if units > 1:
                self.stock_in_box_hint.setText(f"× 每盒 {units} 個 = {v * units} 個")
            else:
                self.stock_in_box_hint.setText("")
        self.stock_in_qty.valueChanged.connect(update_hint)

        self.stock_in_qty.setFocus()

    def _parse_exp_date(self, raw, fmt):
        """將效期字串轉換為 YYYY-MM-DD"""
        try:
            raw = raw.replace('/', '').replace('-', '')
            if fmt in ('YYYYMMDD',):
                return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"
            elif fmt in ('YYMMDD',):
                return f"20{raw[0:2]}-{raw[2:4]}-{raw[4:6]}"
            elif fmt in ('MMDDYYYY',):
                return f"{raw[4:8]}-{raw[0:2]}-{raw[2:4]}"
            elif fmt in ('DDMMYYYY',):
                return f"{raw[4:8]}-{raw[2:4]}-{raw[0:2]}"
            else:
                return raw
        except Exception:
            return raw

    def do_stock_in(self):
        try:
            if not self.stock_in_reagent_id:
                QMessageBox.warning(self, "錯誤", "請先掃描條碼")
                return

            lot_number  = self.stock_in_lot.text().strip()
            expiry_date = self.stock_in_expiry.text().strip()
            box_count   = self.stock_in_qty.value()
            po_number   = self.stock_in_po.text().strip()
            remark      = self.stock_in_remark.text().strip()

            if not lot_number:
                QMessageBox.warning(self, "錯誤", "批號不能為空")
                return
            if not expiry_date:
                QMessageBox.warning(self, "錯誤", "效期不能為空")
                return

            # ===== 依「每盒數量」換算實際入庫數量 =====
            from models import StockIn as SI, ReagentMaster as RM
            reagent_obj   = self.session.query(RM).filter_by(
                id=self.stock_in_reagent_id).first()
            units_per_box = (reagent_obj.units_per_box
                             if reagent_obj and reagent_obj.units_per_box else 1)
            quantity = box_count * units_per_box

            # ===== 偵測是否為新批號 =====
            last = (self.session.query(SI)
                    .filter_by(reagent_id=self.stock_in_reagent_id)
                    .order_by(SI.stock_in_date.desc())
                    .first())
            is_new_lot = (last is None) or (last.lot_number != lot_number)

            success, msg = self.stock_service.stock_in(
                reagent_id=self.stock_in_reagent_id,
                lot_number=lot_number,
                expiry_date=expiry_date,
                quantity=quantity,
                po_number=po_number,
                handler_id=self.user_id,
                remark=remark
            )

            if success:
                # msg 回傳多個編號，以逗號分隔
                codes_part   = msg.split('(編號: ')[-1].rstrip(')')
                in_house_codes = [c.strip() for c in codes_part.split(',')]
                reagent_name   = self.stock_in_reagent_name.text()

                for code in in_house_codes:
                    self._add_stock_in_history_row(
                        datetime.now().strftime('%H:%M:%S'),
                        reagent_name, lot_number, expiry_date, 1, code
                    )
                    self._print_label(
                        reagent_name=reagent_name,
                        lot_number=lot_number,
                        expiry_date=expiry_date,
                        quantity=1,
                        in_house_code=code,
                        label_type='in',
                        handler_name=self.user_real_name,
                        stock_in_date=datetime.now().strftime('%Y-%m-%d'),
                        is_new_lot=is_new_lot
                    )

                self.clear_stock_in_form()
                self.barcode_input.setFocus()
                self.refresh_dashboard()
            else:
                QMessageBox.critical(self, "入庫失敗", msg)

        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def _add_stock_in_history_row(self, time, name, lot, exp, qty, code):
        row = self.stock_in_history.rowCount()
        self.stock_in_history.insertRow(row)
        for col, val in enumerate([time, name, lot, exp, str(qty), code]):
            self.stock_in_history.setItem(row, col, QTableWidgetItem(val))
        # 補列印按鈕
        btn = QPushButton("列印")
        btn.setStyleSheet("background-color: #0078D4; color: white; border-radius: 3px;")
        btn.clicked.connect(lambda _, n=name, l=lot, e=exp, q=qty, c=code:
                            self._print_label(n, l, e, q, c, label_type='in',
                                              handler_name=self.user_real_name))
        self.stock_in_history.setCellWidget(row, 6, btn)

    def _print_label(self, reagent_name, lot_number, expiry_date, quantity, in_house_code,
                     stock_in_date='', label_type='in', handler_name='', is_new_lot=False):
        try:
            import os, subprocess
            from models import SystemSettings
            setting = self.session.query(SystemSettings).filter_by(key='label_printer').first()
            printer_name = setting.value if setting else None

            if label_type == 'in':
                label_path = self._generate_in_label(
                    in_house_code,
                    reagent_name=reagent_name,
                    handler_name=handler_name if handler_name else self.user_real_name,
                    stock_in_date=stock_in_date,
                    is_new_lot=is_new_lot
                )
            else:
                # 出庫標籤：文字資訊
                label_path = self._generate_out_label(
                    reagent_name, expiry_date, in_house_code,
                    handler_name=handler_name)

            if not label_path or not os.path.exists(label_path):
                QMessageBox.warning(self, "列印失敗", f"標籤未產生\n路徑：{label_path}")
                return

            if printer_name:
                subprocess.Popen(['mspaint', '/pt', label_path, printer_name], shell=False)
            else:
                subprocess.Popen(['mspaint', '/p', label_path], shell=False)

        except Exception as e:
            QMessageBox.warning(self, "列印失敗", str(e))

    def _generate_in_label(self, in_house_code, reagent_name='', handler_name='',
                            stock_in_date='', is_new_lot=False):
        try:
            import os
            from utils import LABEL_DIR
            from PIL import Image, ImageDraw, ImageFont, ImageOps

            filename = str(LABEL_DIR / f'in_label_{in_house_code}.png')

            # ===== QR Code 生成 =====
            qr_path = BarcodeGenerator.generate_qrcode(in_house_code)
            if not qr_path or not os.path.exists(qr_path):
                return None

            qr = Image.open(qr_path)

            # 自動切除白邊
            inverted_image = ImageOps.invert(qr.convert('RGB'))
            bbox = inverted_image.getbbox()
            if bbox:
                qr = qr.crop(bbox)

            # ===== 尺寸設定 =====
            DPI = 300
            width        = int(70 * DPI / 25.4)
            height       = int(30 * DPI / 25.4)
            left_margin  = int(20 * DPI / 25.4)
            usable_width = width - left_margin

            # ===== 建立畫布 =====
            img  = Image.new('RGB', (width, height), 'white')
            draw = ImageDraw.Draw(img)

            # ===== 字型 =====
            try:
                font_code    = ImageFont.truetype("arial.ttf", 30)
                font_info    = ImageFont.truetype("arial.ttf", 28)
                font_sm      = ImageFont.truetype("msjh.ttc",  28)
                font_new_lot = ImageFont.truetype("arial.ttf", 36)
            except Exception:
                font_code    = ImageFont.load_default()
                font_info    = font_code
                font_sm      = font_code
                font_new_lot = font_code

            # ===== QR Code 縮小至原本 0.5 倍 =====
            qr_size = int(230 * 0.5)        # 115 px
            qr      = qr.resize((qr_size, qr_size))
            qr_x    = left_margin + (usable_width - qr_size) // 2 + 10
            qr_y    = 10
            img.paste(qr, (qr_x, qr_y))

            # ===== 院內編號（QR 下方）=====
            bbox_t     = draw.textbbox((0, 0), in_house_code, font=font_code)
            text_width = bbox_t[2] - bbox_t[0]
            text_x     = left_margin + (usable_width - text_width) // 2
            text_y     = qr_y + qr_size + 5
            draw.text((text_x, text_y), in_house_code, fill='black', font=font_code)

            # ===== NEW LOT 大字（若批號不同）=====
            if is_new_lot:
                nl_text = 'NEW LOT'
                bbox_nl = draw.textbbox((0, 0), nl_text, font=font_new_lot)
                nl_w    = bbox_nl[2] - bbox_nl[0]
                nl_x    = left_margin + (usable_width - nl_w) // 2
                nl_y    = text_y + 38
                draw.rectangle(
                    [nl_x - 6, nl_y - 4, nl_x + nl_w + 6, nl_y + 42],
                    fill='#D13438'
                )
                draw.text((nl_x, nl_y), nl_text, fill='white', font=font_new_lot)

            # ===== 右側文字 =====
            right_x = qr_x + qr_size + 20
            if right_x < width - 5:
                name_line1 = reagent_name[:7]   if reagent_name else ''
                name_line2 = reagent_name[7:14] if len(reagent_name) > 7 else ''
                draw.text((right_x, 10), name_line1, fill='black', font=font_info)
                if name_line2:
                    draw.text((right_x, 60), name_line2, fill='black', font=font_info)
                y_date  = 120 if name_line2 else 70
                in_date = stock_in_date or datetime.now().strftime('%Y-%m-%d')
                draw.text((right_x, y_date),      in_date,      fill='#222222', font=font_sm)
                draw.text((right_x, y_date + 60), handler_name, fill='#444444', font=font_sm)

            img.save(filename)
            return filename

        except Exception as e:
            print(f"入庫標籤生成失敗: {e}")
            return None
        
    def _generate_out_label(self, reagent_name, expiry_date, in_house_code,
                            handler_name=''):
        """出庫標籤：試劑名稱、出庫日期、有效期限、出庫人名稱"""
        try:
            import os
            from datetime import datetime  # 確保有引入 datetime
            from utils import LABEL_DIR
            from PIL import Image, ImageDraw, ImageFont
            
            out_date = datetime.now().strftime('%Y-%m-%d')
            filename = str(LABEL_DIR / f'out_label_{in_house_code}_{out_date}.png')

            # 70mm x 30mm @ 300 DPI = 827 x 472 px
            DPI = 300
            width = int(70 * DPI / 25.4)
            height = int(30 * DPI / 25.4)
            img = Image.new('RGB', (width, height), 'white')
            draw = ImageDraw.Draw(img)

            # ===== 🌟 字型設定修正（更換為支援繁體中文的微軟正黑體） =====
            try:
                # 優先嘗試系統內建簡寫
                font_path = "msjh.ttc" 
                font_title = ImageFont.truetype(font_path, 36)
                font_body  = ImageFont.truetype(font_path, 30)
                font_sm    = ImageFont.truetype(font_path, 24)
            except Exception:
                try:
                    # 如果上面失敗，嘗試 Windows 標準絕對路徑
                    font_path = "C:/Windows/Fonts/msjh.ttc"
                    font_title = ImageFont.truetype(font_path, 50)
                    font_body  = ImageFont.truetype(font_path, 40)
                    font_sm    = ImageFont.truetype(font_path, 24)
                except Exception:
                    # Linux 或 Mac 環境若失敗，可在此處替換對應字型路徑
                    print("警告：找不到微軟正黑體，將使用預設字型（中文會無法顯示）")
                    font_title = ImageFont.load_default()
                    font_body  = font_title
                    font_sm    = font_title

            # ===== 🌟 新增：統一控制所有文字的左邊距 (X軸) =====
            # 原本是 100，現在改為 180。如果印出來覺得還不夠右邊，可以自己改成 200 或 220
            start_x = 500

            # ===== 繪製文字與線條 =====
            # 將原本的 100 全部替換為 start_x
            draw.text((start_x, 0), "衛生福利部花蓮醫院", fill='black', font=font_body)
           
            draw.text((start_x, 35),  f"試劑：{reagent_name[:20]}", fill='black', font=font_title)
            draw.text((start_x, 95), f"效期：{expiry_date}",        fill='black', font=font_body)
            draw.text((start_x, 155), f"出庫日期：{out_date}",           fill='black', font=font_body)
            draw.text((start_x, 225), f"人員：{handler_name}",       fill='black', font=font_body)

            img.save(filename)
            return filename
        except Exception as e:
            print(f"出庫標籤生成失敗: {e}")
            return None
        
    def _generate_out_label_old(self, reagent_name, lot_number, expiry_date,
                             stock_in_date, in_house_code, qrcode_path):
        pass  # 舊版本已棄用

    def clear_stock_in_form(self):
        self.barcode_input.clear()
        self.stock_in_reagent_name.clear()
        self.stock_in_lot.clear()
        self.stock_in_expiry.clear()
        self.stock_in_qty.setValue(1)
        self.stock_in_po.clear()
        self.stock_in_remark.clear()
        self.stock_in_reagent_id = None
        self.stock_in_box_hint.setText("")
        try:
            self.stock_in_qty.valueChanged.disconnect()
        except Exception:
            pass
    
    def open_inventory_search_dialog(self):
        """開啟庫存搜尋視窗，選擇後帶入出庫院內編號"""
        from models import Inventory, ReagentMaster

        dialog = QDialog(self)
        dialog.setWindowTitle("搜尋庫存")
        dialog.resize(1000, 600)
        layout = QVBoxLayout(dialog)

        # ── 搜尋條件列 ──
        cond_layout = QHBoxLayout()

        cond_layout.addWidget(QLabel("試劑名稱:"))
        f_name = QLineEdit()
        f_name.setMinimumHeight(34)
        f_name.setMinimumWidth(150)
        cond_layout.addWidget(f_name)

        cond_layout.addWidget(QLabel("LOT號:"))
        f_lot = QLineEdit()
        f_lot.setMinimumHeight(34)
        f_lot.setMinimumWidth(120)
        cond_layout.addWidget(f_lot)

        cond_layout.addWidget(QLabel("院內編號:"))
        f_code = QLineEdit()
        f_code.setMinimumHeight(34)
        f_code.setMinimumWidth(140)
        cond_layout.addWidget(f_code)

        cond_layout.addWidget(QLabel("廠牌:"))
        f_brand = QLineEdit()
        f_brand.setMinimumHeight(34)
        f_brand.setMinimumWidth(100)
        cond_layout.addWidget(f_brand)

        show_zero_chk = QCheckBox("顯示庫存為0")
        cond_layout.addWidget(show_zero_chk)

        search_btn2 = QPushButton("搜尋")
        search_btn2.setMinimumHeight(34)
        search_btn2.setStyleSheet("background:#0078D4; color:white; border-radius:4px;")
        cond_layout.addWidget(search_btn2)

        clear_btn2 = QPushButton("清除")
        clear_btn2.setMinimumHeight(34)
        cond_layout.addWidget(clear_btn2)

        cond_layout.addStretch()
        layout.addLayout(cond_layout)

        count_lbl = QLabel("共 0 筆")
        layout.addWidget(count_lbl)

        # ── 結果表格 ──
        table = QTableWidget()
        table.setColumnCount(8)
        table.setHorizontalHeaderLabels([
            "試劑名稱", "廠牌", "LOT號", "院內編號",
            "有效期限", "現有庫存", "狀態", "安全庫存"
        ])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(table)

        hint = QLabel("雙擊或選取後按「確認選擇」帶入院內編號")
        hint.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(hint)

        # ── 底部按鈕 ──
        bottom = QHBoxLayout()
        confirm_btn = QPushButton("確認選擇")
        confirm_btn.setMinimumHeight(40)
        confirm_btn.setStyleSheet("background:#107C10; color:white; font-weight:bold; border-radius:4px;")
        cancel_btn2 = QPushButton("取消")
        cancel_btn2.setMinimumHeight(40)
        cancel_btn2.clicked.connect(dialog.reject)
        bottom.addWidget(confirm_btn)
        bottom.addWidget(cancel_btn2)
        bottom.addStretch()
        layout.addLayout(bottom)

        def do_search():
            table.setRowCount(0)
            q = (self.session.query(Inventory)
                 .join(ReagentMaster, Inventory.reagent_id == ReagentMaster.id))
            if not show_zero_chk.isChecked():
                q = q.filter(Inventory.current_quantity > 0)
            if f_name.text().strip():
                q = q.filter(ReagentMaster.reagent_name.contains(f_name.text().strip()))
            if f_lot.text().strip():
                q = q.filter(Inventory.lot_number.contains(f_lot.text().strip()))
            if f_code.text().strip():
                q = q.filter(Inventory.in_house_code.contains(f_code.text().strip()))
            if f_brand.text().strip():
                q = q.filter(ReagentMaster.brand.contains(f_brand.text().strip()))
            items = q.order_by(
                ReagentMaster.reagent_name,
                Inventory.expiry_date
            ).all()

            for idx, inv in enumerate(items):
                table.insertRow(idx)
                status, color = DateHelper.get_expiry_status(inv.expiry_date or '')
                for col, v in enumerate([
                    inv.reagent.reagent_name,
                    inv.reagent.brand or '',
                    inv.lot_number or '',
                    inv.in_house_code or '',
                    inv.expiry_date or '',
                    str(inv.current_quantity),
                    status,
                    str(inv.reagent.safety_stock)
                ]):
                    cell = QTableWidgetItem(v)
                    if col == 6:
                        cell.setBackground(QColor(color))
                        cell.setForeground(QColor('white'))
                    elif col == 5 and inv.current_quantity <= inv.reagent.safety_stock:
                        cell.setBackground(QColor('#FFB900'))
                    table.setItem(idx, col, cell)
            count_lbl.setText(f"共 {len(items)} 筆")

        def on_confirm():
            rows = table.selectionModel().selectedRows()
            if not rows:
                QMessageBox.warning(dialog, "提示", "請先選取一筆庫存")
                return
            in_house_code = table.item(rows[0].row(), 3).text()
            if not in_house_code:
                QMessageBox.warning(dialog, "提示", "該筆庫存沒有院內編號")
                return
            self.stock_out_barcode.setText(in_house_code)
            self.scan_stock_out_barcode()
            dialog.accept()

        search_btn2.clicked.connect(do_search)
        confirm_btn.clicked.connect(on_confirm)
        table.doubleClicked.connect(on_confirm)
        clear_btn2.clicked.connect(lambda: [
            f_name.clear(), f_lot.clear(), f_code.clear(), f_brand.clear()
        ])
        f_name.returnPressed.connect(do_search)
        f_lot.returnPressed.connect(do_search)
        f_code.returnPressed.connect(do_search)
        show_zero_chk.stateChanged.connect(lambda _: do_search())

        do_search()
        dialog.exec()

    def scan_stock_out_barcode(self):
        in_house_code = self.stock_out_barcode.text().strip()
        if not in_house_code:
            return
        try:
            from models import StockIn, StockOut
            from sqlalchemy import func
            stock_in = self.session.query(StockIn).filter_by(
                in_house_code=in_house_code
            ).first()
            if stock_in:
                self._current_stock_in = stock_in
                self.stock_out_name.setText(stock_in.reagent.reagent_name)
                self.stock_out_lot.setText(stock_in.lot_number)
                self.stock_out_expiry.setText(stock_in.expiry_date)
                used = self.session.query(func.sum(StockOut.quantity)).filter_by(
                    stock_in_id=stock_in.id).scalar() or 0
                available = stock_in.quantity - used
                self.stock_out_available.setText(str(available))
                self.stock_out_qty.setMaximum(available)
                self.stock_out_qty.setValue(1)
                self.stock_out_qty.setFocus()
            else:
                QMessageBox.warning(self, "未找到", "院內編號不存在")
                self.stock_out_barcode.clear()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def do_stock_out(self):
        try:
            in_house_code = self.stock_out_barcode.text().strip()
            if not in_house_code:
                QMessageBox.warning(self, "錯誤", "請掃描院內編號")
                return
            quantity = self.stock_out_qty.value()

            # ===== 偵測批號是否與前一次出庫不同 =====
            lot_now = self.stock_out_lot.text().strip()
            if self._current_stock_in:
                from models import StockOut as SO, StockIn as SI
                last_out = (self.session.query(SO)
                            .join(SI, SO.stock_in_id == SI.id)
                            .filter(SI.reagent_id == self._current_stock_in.reagent_id)
                            .order_by(SO.stock_out_date.desc())
                            .first())
                if last_out and last_out.stock_in.lot_number != lot_now:
                    reply = QMessageBox.warning(
                        self, "批號變更",
                        f"⚠️ 注意：批號已變更！\n\n"
                        f"上次出庫批號：{last_out.stock_in.lot_number}\n"
                        f"本次出庫批號：{lot_now}\n\n"
                        f"確定要繼續出庫？",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    if reply == QMessageBox.StandardButton.No:
                        return

            success, msg = self.stock_service.stock_out(
                in_house_code=in_house_code,
                quantity=quantity,
                usage_department=self.stock_out_dept.text(),
                usage_equipment=self.stock_out_equipment.text(),
                handler_id=self.user_id,
                remark=self.stock_out_remark.text() if hasattr(self.stock_out_remark, 'text') else self.stock_out_remark.toPlainText()
            )
            if success:
                name    = self.stock_out_name.text()
                lot     = self.stock_out_lot.text()
                exp     = self.stock_out_expiry.text()
                handler = self.user_real_name

                row = self.stock_out_history.rowCount()
                self.stock_out_history.insertRow(row)
                for col, val in enumerate([
                    datetime.now().strftime('%H:%M:%S'),
                    name, lot, exp, str(quantity), in_house_code
                ]):
                    self.stock_out_history.setItem(row, col, QTableWidgetItem(val))
                btn = QPushButton("列印")
                btn.setStyleSheet("background-color: #0078D4; color: white; border-radius: 3px;")
                btn.clicked.connect(lambda _, n=name, e=exp, c=in_house_code, h=handler:
                                    self._print_label(n, '', e, 0, c,
                                                      label_type='out', handler_name=h))
                self.stock_out_history.setCellWidget(row, 6, btn)

                # 自動列印出庫標籤
                self._print_label(name, '', exp, quantity, in_house_code,
                                  label_type='out', handler_name=handler)

                # 若試劑需驗收且批次未完成，額外印驗收標籤
                if self._current_stock_in:
                    sin = self._current_stock_in
                    sin_date = sin.stock_in_date.strftime('%Y-%m-%d')
                    if self.qc_service.is_qc_required_for_out(
                            sin.reagent_id, sin.lot_number, sin_date):
                        qc_path = self._generate_qc_label(
                            name,
                            datetime.now().strftime('%Y-%m-%d'),
                            in_house_code
                        )
                        if qc_path:
                            import os, subprocess
                            from models import SystemSettings
                            setting = self.session.query(SystemSettings).filter_by(
                                key='label_printer').first()
                            printer_name = setting.value if setting else None
                            if printer_name:
                                subprocess.Popen(
                                    ['mspaint', '/pt', qc_path, printer_name], shell=False)
                            else:
                                subprocess.Popen(['mspaint', '/p', qc_path], shell=False)

                self.clear_stock_out_form()
                self.refresh_dashboard()
            else:
                QMessageBox.critical(self, "出庫失敗", msg)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))

    def do_scrap(self):
        """執行報廢"""
        try:
            reagent_id = self.scrap_reagent.currentData()
            lot_number = self.scrap_lot.text().strip()
            quantity = self.scrap_qty.value()
            reason = self.scrap_reason.currentText()
            remark = self.scrap_remark.toPlainText()
            
            if not lot_number:
                QMessageBox.warning(self, "輸入錯誤", "請輸入LOT號")
                return
            
            success, msg = self.scrap_service.create_scrap_record(
                reagent_id=reagent_id,
                lot_number=lot_number,
                quantity=quantity,
                reason=reason,
                remark=remark
            )
            
            QMessageBox.information(self, "結果", msg)
            if success:
                self.clear_scrap_form()
                self.refresh_dashboard()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    def do_traceability_query(self):
        """執行追溯查詢"""
        try:
            in_house_code = self.trace_code.text().strip()
            lot_number = self.trace_lot.text().strip()
            
            if not in_house_code and not lot_number:
                QMessageBox.warning(self, "輸入錯誤", "請輸入院內編號或LOT號")
                return
            
            result = self.traceability_service.get_traceability(
                in_house_code=in_house_code if in_house_code else None,
                lot_number=lot_number if lot_number else None
            )
            
            # 格式化結果顯示
            output = "=== 入庫紀錄 ===\n"
            for record in result['stock_in_records']:
                output += f"日期: {record['date']}, 數量: {record['quantity']}, 採購單: {record['po_number']}, 經手: {record['handler']}\n"
            
            output += "\n=== 出庫紀錄 ===\n"
            for record in result['stock_out_records']:
                output += f"日期: {record['date']}, 數量: {record['quantity']}, 單位: {record['department']}, 儀器: {record['equipment']}\n"
            
            output += "\n=== 盤點紀錄 ===\n"
            for record in result['inventory_checks']:
                output += f"日期: {record['date']}, 系統: {record['system_qty']}, 實際: {record['actual_qty']}, 差異: {record['difference']}\n"
            
            output += "\n=== 報廢紀錄 ===\n"
            for record in result['scrap_records']:
                output += f"日期: {record['date']}, 數量: {record['quantity']}, 原因: {record['reason']}\n"
            
            self.trace_result.setText(output)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    def export_excel_report(self):
        """匯出Excel報表"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存Excel報表", "", "Excel Files (*.xlsx)"
            )
            if not file_path:
                return
            
            report_type = self.report_type.currentText()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report_type
            
            # 標題
            ws['A1'] = "衛生福利部花蓮醫院"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A2'] = f"試劑管理系統 - {report_type}"
            ws['A2'].font = Font(size=12, bold=True)
            ws['A3'] = f"產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            # 根據報表類型添加內容
            if report_type == "庫存報表":
                headers = ["試劑名稱", "LOT號", "有效期限", "院內編號", "現有庫存", "安全庫存", "狀態"]
                ws.append(headers)
                
                inventory_items = self.stock_service.get_current_inventory()
                for item in inventory_items:
                    status, _ = DateHelper.get_expiry_status(item.expiry_date)
                    ws.append([
                        item.reagent.reagent_name,
                        item.lot_number,
                        item.expiry_date,
                        item.in_house_code,
                        item.current_quantity,
                        item.reagent.safety_stock,
                        status
                    ])
            
            wb.save(file_path)
            QMessageBox.information(self, "成功", f"報表已匯出：{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    def export_pdf_report(self):
        """匯出PDF報表（使用支援中文的 _export_pdf_table）"""
        try:
            inventory_items = self.stock_service.get_current_inventory()
            if not inventory_items:
                QMessageBox.warning(self, "提示", "沒有資料可匯出")
                return

            headers = ["試劑名稱", "LOT號", "現有庫存", "有效期限", "狀態"]
            data_rows = []
            for item in inventory_items:
                status, _ = DateHelper.get_expiry_status(item.expiry_date)
                data_rows.append([
                    item.reagent.reagent_name,
                    item.lot_number,
                    str(item.current_quantity),
                    item.expiry_date,
                    status
                ])

            title = f"試劑管理報表  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            self._export_pdf_table(title, headers, data_rows)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    def add_user_dialog(self):
        """新增使用者對話框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("新增使用者")
        dialog.setGeometry(200, 200, 400, 300)
        
        layout = QVBoxLayout()
        
        layout.addWidget(QLabel("使用者名稱:"))
        username = QLineEdit()
        layout.addWidget(username)
        
        layout.addWidget(QLabel("密碼:"))
        password = QLineEdit()
        password.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addWidget(password)
        
        layout.addWidget(QLabel("真實姓名:"))
        real_name = QLineEdit()
        layout.addWidget(real_name)
        
        layout.addWidget(QLabel("角色:"))
        role = QComboBox()
        role.addItems(["user", "admin"])
        layout.addWidget(role)
        
        layout.addWidget(QLabel("部門:"))
        dept = QLineEdit()
        layout.addWidget(dept)
        
        btn_layout = QHBoxLayout()
        
        def save_user():
            try:
                success, msg = self.user_service.create_user(
                    username=username.text(),
                    password=password.text(),
                    real_name=real_name.text(),
                    role=role.currentText(),
                    department=dept.text()
                )
                QMessageBox.information(self, "完成", msg)
                if success:
                    self.refresh_user_table()
                    dialog.accept()
            except Exception as e:
                QMessageBox.critical(self, "錯誤", str(e))
        
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(save_user)
        btn_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        dialog.setLayout(layout)
        dialog.exec()
    
    def confirm_check(self, row_idx, item_id):
        """確認盤點"""
        try:
            actual_qty_widget = self.check_table.cellWidget(row_idx, 3)
            actual_qty = actual_qty_widget.value()
            
            # 從第2列（系統庫存）獲取值
            system_qty = int(self.check_table.item(row_idx, 2).text())
            difference = actual_qty - system_qty
            
            # 更新差異顯示
            diff_item = self.check_table.item(row_idx, 4)
            diff_item.setText(str(difference))
            
            # 保存到資料庫
            reagent_name = self.check_table.item(row_idx, 0).text()
            lot_number = self.check_table.item(row_idx, 1).text()
            
            # 找到對應的試劑ID
            inventory_items = self.stock_service.get_current_inventory()
            for item in inventory_items:
                if item.reagent.reagent_name == reagent_name and item.lot_number == lot_number:
                    success, msg = self.check_service.create_check_record(
                        reagent_id=item.reagent_id,
                        lot_number=lot_number,
                        actual_quantity=actual_qty,
                        remark=f'系統庫存：{system_qty}'
                    )
                    QMessageBox.information(self, "結果", msg)
                    break
        except Exception as e:
            QMessageBox.critical(self, "錯誤", str(e))
    
    # ============ 清除表單方法 ============
    
    def clear_stock_out_form(self):
        self.stock_out_barcode.clear()
        self.stock_out_name.clear()
        self.stock_out_lot.clear()
        self.stock_out_expiry.clear()
        self.stock_out_available.clear()
        self.stock_out_qty.setValue(1)
        self.stock_out_dept.clear()
        self.stock_out_equipment.clear()
        self.stock_out_remark.clear()
        self._current_stock_in = None
        self.stock_out_barcode.setFocus()
    
    def clear_scrap_form(self):
        """清除報廢表單"""
        self.scrap_lot.clear()
        self.scrap_qty.setValue(1)
        self.scrap_remark.clear()
        self.scrap_lot.setFocus()
    
    def clear_query_conditions(self):
        self.query_name.clear()
        self.query_lot.clear()
        self.query_date_from.setDate(QDate.currentDate().addDays(-30))
        self.query_date_to.setDate(QDate.currentDate())
        self.inventory_table.setRowCount(0)
        self.query_count_label.setText("共 0 筆")
    
    def logout(self):
        """登出"""
        reply = QMessageBox.question(self, "確認", "確定要登出嗎？", 
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.session.close()
            self.close()


if __name__ == '__main__':
    from ui_login import LoginWindow
    import sys
    
    app = QApplication(sys.argv)
    
    # 顯示登入視窗
    login_window = LoginWindow()
    
    def on_login_success(user):
        main_window = MainWindow(user)
        main_window.show()
        login_window.hide()
    
    login_window.login_success.connect(on_login_success)
    login_window.exec()
